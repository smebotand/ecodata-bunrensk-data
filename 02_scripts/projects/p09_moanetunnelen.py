"""
Project 09 - Moanetunnelen Extraction Script

Extracts data from the Innsyn PDF to populate the 5-table data schema:
- samples.csv
- results.csv
- classifications.csv
- decisions.csv
- extraction_summary.csv

Source pages:
- Pages 2-5: SVV notat with sample info, classifications, and decisions
- Pages 6-30: ALS lab reports with detailed chemical analysis results
"""

import sys
import re
from pathlib import Path
import pandas as pd
from datetime import datetime

# Add parent to path for lib imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from lib.pdf_utils import extract_text, extract_pages
from lib.chemistry import normalize_parameter_name, parse_value, TILSTANDSKLASSER
from lib.export import save_to_csv

# ============================================================
# PROJECT CONFIGURATION
# ============================================================

PROJECT_CODE = '09_moanetunnelen'
PROJECT_NAME = 'Moanetunnelen'
TUNNEL_NAME = 'Moanetunnelen'

BASE_DIR = Path(__file__).parent.parent.parent
INBOX_DIR = BASE_DIR / '00_inbox' / 'SVV' / 'Moanetunnelen'
OUTPUT_DIR = BASE_DIR / '01_projects' / '09_moanetunnelen' / 'extracted'

PDF_FILE = 'Innsyn 25-339107-1Svar på innsyn - Innsynskrav i - 16171864-58 med vedlegg (1).pdf'

# Page ranges
NOTAT_PAGES = list(range(2, 6))      # Pages 2-5: SVV notat
LAB_REPORT_PAGES = list(range(6, 31)) # Pages 6-30: ALS lab reports

# ============================================================
# SAMPLE METADATA (extracted from pages 2-5 of the notat)
# ============================================================

SAMPLES = [
    # First sampling 20.09.2017
    {'sample_key': '1', 'sample_id': 'p09-MOA-001', 'sample_date': '2017-09-20', 
     'location_type': 'vegbane', 'profile_start': 9850, 'profile_end': 9950,
     'sample_type': 'bunnrensk', 'lab_reference': 'N00529080'},
    {'sample_key': '2', 'sample_id': 'p09-MOA-002', 'sample_date': '2017-09-20',
     'location_type': 'vegbane', 'profile_start': 9950, 'profile_end': 10100,
     'sample_type': 'bunnrensk', 'lab_reference': 'N00529081'},
    {'sample_key': '3', 'sample_id': 'p09-MOA-003', 'sample_date': '2017-09-20',
     'location_type': 'grøft', 'profile_start': 9950, 'profile_end': 10100,
     'sample_type': 'bunnrensk', 'lab_reference': 'N00529082'},
    {'sample_key': '4', 'sample_id': 'p09-MOA-004', 'sample_date': '2017-09-20',
     'location_type': 'grøft', 'profile_start': 9850, 'profile_end': 9950,
     'sample_type': 'bunnrensk', 'lab_reference': 'N00529083'},
    # Second sampling 28.09.2017
    {'sample_key': 'T1', 'sample_id': 'p09-MOA-T1', 'sample_date': '2017-09-28',
     'location_type': 'vegbane', 'profile_start': 9850, 'profile_end': 9900,
     'sample_type': 'bunnrensk', 'lab_reference': 'N00531434'},
    {'sample_key': 'T2', 'sample_id': 'p09-MOA-T2', 'sample_date': '2017-09-28',
     'location_type': 'vegbane', 'profile_start': 9900, 'profile_end': 9950,
     'sample_type': 'bunnrensk', 'lab_reference': 'N00531435'},
    {'sample_key': 'D1', 'sample_id': 'p09-MOA-D1', 'sample_date': '2017-09-28',
     'location_type': 'deponi', 'profile_start': None, 'profile_end': None,
     'sample_type': 'blandeprøve', 'lab_reference': 'N00531432'},
    {'sample_key': 'D2', 'sample_id': 'p09-MOA-D2', 'sample_date': '2017-09-28',
     'location_type': 'deponi', 'profile_start': None, 'profile_end': None,
     'sample_type': 'blandeprøve', 'lab_reference': 'N00531433'},
]

# Classifications from Tabell 1 and 2 in the notat (pages 2-3)
# limiting_parameters is a list to support multiple limiting factors
CLASSIFICATIONS = [
    {'sample_key': '1', 'tilstandsklasse': 4, 'limiting_parameters': ['As']},
    {'sample_key': '2', 'tilstandsklasse': 3, 'limiting_parameters': ['Cu']},
    {'sample_key': '3', 'tilstandsklasse': 2, 'limiting_parameters': []},
    {'sample_key': '4', 'tilstandsklasse': 4, 'limiting_parameters': ['THC_C12-C35']},
    {'sample_key': 'T1', 'tilstandsklasse': 3, 'limiting_parameters': ['THC', 'Cu']},
    {'sample_key': 'T2', 'tilstandsklasse': 4, 'limiting_parameters': ['THC']},
    {'sample_key': 'D1', 'tilstandsklasse': 3, 'limiting_parameters': ['THC', 'Cu']},
    {'sample_key': 'D2', 'tilstandsklasse': 3, 'limiting_parameters': ['THC']},
]

# Decisions from the notat (pages 2-5)
DECISIONS = [
    {'sample_key': '1', 'decision': 'Nye prøver tas for å se på arsen-nivået', 
     'destination': 'avventer', 'notes': 'Tilstandsklasse IV for arsen'},
    {'sample_key': '2', 'decision': 'Kan disponeres som planlagt', 
     'destination': 'gjenbruk', 'notes': 'Lett/moderat forurenset'},
    {'sample_key': '3', 'decision': 'Blandet med prøve 4, nye prøver tas', 
     'destination': 'avventer', 'notes': ''},
    {'sample_key': '4', 'decision': 'Må kjøres til godkjent deponi', 
     'destination': 'deponi', 'notes': 'Oljeutslipp er entreprenørens ansvar'},
    {'sample_key': 'T1', 'decision': 'Kan benyttes i anlegget', 
     'destination': 'gjenbruk', 'notes': 'Profil 9850-9900'},
    {'sample_key': 'T2', 'decision': 'Må kjøres til godkjent deponi', 
     'destination': 'deponi', 'notes': 'Oljeutslipp - profil 9900-9950'},
    {'sample_key': 'D1', 'decision': 'Kan ikke benyttes i anlegget', 
     'destination': 'deponi', 'notes': 'Grøftemasser på deponi'},
    {'sample_key': 'D2', 'decision': 'Kan ikke benyttes i anlegget', 
     'destination': 'deponi', 'notes': 'Grøftemasser på deponi'},
]

# Mapping sample_key to sample_id
SAMPLE_KEY_TO_ID = {s['sample_key']: s['sample_id'] for s in SAMPLES}


def parse_lab_results(text: str) -> list:
    """
    Parse ALS lab report format from PDF text (pages 6-30).
    
    The lab reports have format:
    Deresprøvenavn X
    Sediment
    Labnummer NXXXXXXXX
    Analyse Resultater Usikkerhet (±) Enhet Metode Utført Sign
    ParameterNameaulev value uncertainty unit ...
    
    Note: PDF extraction often strips spaces, so e.g. "As(Arsen)aulev 51.9 10.4 mg/kgTS"
    """
    results = []
    
    # Split by sample sections - find "Deresprøvenavn"
    pattern = r'Deresprøvenavn\s+(\w+)'
    
    # Find all sample sections
    sample_matches = list(re.finditer(pattern, text))
    
    for i, match in enumerate(sample_matches):
        sample_key = match.group(1)
        start_pos = match.end()
        
        # End position is start of next sample or end of text
        if i + 1 < len(sample_matches):
            end_pos = sample_matches[i + 1].start()
        else:
            end_pos = len(text)
        
        section = text[start_pos:end_pos]
        
        # Get lab number
        lab_match = re.search(r'Labnummer\s+(\w+)', section)
        lab_num = lab_match.group(1) if lab_match else ''
        
        # Get sample_id from mapping
        sample_id = SAMPLE_KEY_TO_ID.get(sample_key, f'p09-MOA-{sample_key}')
        
        # Parse all analysis lines using a generic line-by-line approach
        # Format: ParameterName[suffix]aulev <value> [uncertainty] mg/kgTS ...
        # Note: PDF strips spaces, so we match patterns carefully
        
        section_results = parse_section_results(section, sample_id)
        results.extend(section_results)
    
    return results


def parse_section_results(section: str, sample_id: str) -> list:
    """
    Parse all parameter results from a sample section.
    Uses line-by-line parsing after cleaning footnote references.
    """
    results = []
    found_params = set()  # Track what we've found to avoid duplicates
    
    # =========================================================================
    # CLEAN THE TEXT - Remove footnote references like "aulev", "^aulev", "a ulev"
    # =========================================================================
    clean_section = section
    clean_section = re.sub(r'\^?a\s*u\s*l\s*e\s*v', ' ', clean_section)  # Remove aulev variants
    clean_section = re.sub(r'\^', '', clean_section)  # Remove caret symbols
    clean_section = re.sub(r'\s+', ' ', clean_section)  # Normalize whitespace
    
    # =========================================================================
    # PARAMETER MAPPING: Norwegian name pattern -> (param_code, display_name)
    # =========================================================================
    PARAM_MAP = {
        # Dry matter (%)
        r'Tørrstoff\s*\([ED]\)': ('DryMatter', 'Tørrstoff', '%'),
        
        # METALS
        r'As\s*\(Arsen\)': ('As', 'Arsen', 'mg/kg'),
        r'Cd\s*\(Kadmium\)': ('Cd', 'Kadmium', 'mg/kg'),
        r'Cr\s*\(Krom\)': ('Cr', 'Krom', 'mg/kg'),
        r'Cu\s*\(Kopper\)': ('Cu', 'Kopper', 'mg/kg'),
        r'Hg\s*\(Kvikksølv\)': ('Hg', 'Kvikksølv', 'mg/kg'),
        r'Ni\s*\(Nikkel\)': ('Ni', 'Nikkel', 'mg/kg'),
        r'Pb\s*\(Bly\)': ('Pb', 'Bly', 'mg/kg'),
        r'Zn\s*\(Sink\)': ('Zn', 'Sink', 'mg/kg'),
        r'Cr6\+': ('Cr_VI', 'Cr6+', 'mg/kg'),
        
        # PCB CONGENERS
        r'PCB\s*28(?!\d)': ('PCB28', 'PCB 28', 'mg/kg'),
        r'PCB\s*52(?!\d)': ('PCB52', 'PCB 52', 'mg/kg'),
        r'PCB\s*101(?!\d)': ('PCB101', 'PCB 101', 'mg/kg'),
        r'PCB\s*118(?!\d)': ('PCB118', 'PCB 118', 'mg/kg'),
        r'PCB\s*138(?!\d)': ('PCB138', 'PCB 138', 'mg/kg'),
        r'PCB\s*153(?!\d)': ('PCB153', 'PCB 153', 'mg/kg'),
        r'PCB\s*180(?!\d)': ('PCB180', 'PCB 180', 'mg/kg'),
        r'Sum\s*PCB-7': ('PCB7', 'Sum PCB-7', 'mg/kg'),
        
        # PAH - 16 EPA COMPOUNDS
        r'Naftalen(?!-)': ('Naphthalene', 'Naftalen', 'mg/kg'),
        r'Acenaftylen': ('Acenaphthylene', 'Acenaftylen', 'mg/kg'),
        r'Acenaften(?!-)': ('Acenaphthene', 'Acenaften', 'mg/kg'),
        r'Fluoren(?!-)': ('Fluorene', 'Fluoren', 'mg/kg'),
        r'Fenantren': ('Phenanthrene', 'Fenantren', 'mg/kg'),
        r'Antracen(?!-)': ('Anthracene', 'Antracen', 'mg/kg'),
        r'Fluoranten': ('Fluoranthene', 'Fluoranten', 'mg/kg'),
        r'Pyren(?!-)': ('Pyrene', 'Pyren', 'mg/kg'),
        r'Benso\(a\)antracen': ('BaA', 'Benzo(a)antracen', 'mg/kg'),
        r'Krysen': ('Chrysene', 'Krysen', 'mg/kg'),
        r'Benso\(b\)fluoranten': ('BbF', 'Benzo(b)fluoranten', 'mg/kg'),
        r'Benso\(k\)fluoranten': ('BkF', 'Benzo(k)fluoranten', 'mg/kg'),
        r'Benso\(a\)pyren': ('BaP', 'Benzo(a)pyren', 'mg/kg'),
        r'Dibenso\(ah\)antracen': ('DahA', 'Dibenzo(ah)antracen', 'mg/kg'),
        r'Benso\(ghi\)perylen': ('BghiP', 'Benzo(ghi)perylen', 'mg/kg'),
        r'Indeno\(1,?2,?3-?c,?d\)pyren': ('IcdP', 'Indeno(123cd)pyren', 'mg/kg'),
        r'Sum\s*PAH-16': ('PAH16', 'Sum PAH-16', 'mg/kg'),
        
        # BTEX
        r'Bensen(?!o)': ('Benzene', 'Benzen', 'mg/kg'),
        r'Toluen': ('Toluene', 'Toluen', 'mg/kg'),
        r'Etylbensen': ('Ethylbenzene', 'Etylbensen', 'mg/kg'),
        r'Xylener': ('Xylene', 'Xylener', 'mg/kg'),
        r'Sum\s*BTEX': ('BTEX', 'Sum BTEX', 'mg/kg'),
        
        # THC FRACTIONS
        r'Fraksjon\s*>?\s*C5-C6': ('THC_C5-C6', 'Fraksjon >C5-C6', 'mg/kg'),
        r'Fraksjon\s*>?\s*C6-C8': ('THC_C6-C8', 'Fraksjon >C6-C8', 'mg/kg'),
        r'Fraksjon\s*>?\s*C8-C10': ('THC_C8-C10', 'Fraksjon >C8-C10', 'mg/kg'),
        r'Fraksjon\s*>?\s*C10-C12': ('THC_C10-C12', 'Fraksjon >C10-C12', 'mg/kg'),
        r'Fraksjon\s*>?\s*C12-C16': ('THC_C12-C16', 'Fraksjon >C12-C16', 'mg/kg'),
        r'Fraksjon\s*>?\s*C16-C35': ('THC_C16-C35', 'Fraksjon >C16-C35', 'mg/kg'),
        r'Sum\s*>?\s*C12-C35': ('THC_C12-C35', 'Sum >C12-C35', 'mg/kg'),
        
        # CYANIDE
        r'Cyanid-fri': ('CN', 'Cyanid-fri', 'mg/kg'),
        
        # CHLOROPHENOLS
        r'2-Monoklorfenol': ('CP_2-MCP', '2-Monoklorfenol', 'mg/kg'),
        r'3-Monoklorfenol': ('CP_3-MCP', '3-Monoklorfenol', 'mg/kg'),
        r'4-Monoklorfenol': ('CP_4-MCP', '4-Monoklorfenol', 'mg/kg'),
        r'2,3-Diklorfenol': ('CP_2,3-DCP', '2,3-Diklorfenol', 'mg/kg'),
        r'2,4\+2,5-Diklorfenol': ('CP_2,4+2,5-DCP', '2,4+2,5-Diklorfenol', 'mg/kg'),
        r'2,6-Diklorfenol': ('CP_2,6-DCP', '2,6-Diklorfenol', 'mg/kg'),
        r'3,4-Diklorfenol': ('CP_3,4-DCP', '3,4-Diklorfenol', 'mg/kg'),
        r'3,5-Diklorfenol': ('CP_3,5-DCP', '3,5-Diklorfenol', 'mg/kg'),
        r'2,3,4-Triklorfenol': ('CP_2,3,4-TCP', '2,3,4-Triklorfenol', 'mg/kg'),
        r'2,3,5-Triklorfenol': ('CP_2,3,5-TCP', '2,3,5-Triklorfenol', 'mg/kg'),
        r'2,3,6-Triklorfenol': ('CP_2,3,6-TCP', '2,3,6-Triklorfenol', 'mg/kg'),
        r'2,4,5-Triklorfenol': ('CP_2,4,5-TCP', '2,4,5-Triklorfenol', 'mg/kg'),
        r'2,4,6-Triklorfenol': ('CP_2,4,6-TCP', '2,4,6-Triklorfenol', 'mg/kg'),
        r'3,4,5-Triklorfenol': ('CP_3,4,5-TCP', '3,4,5-Triklorfenol', 'mg/kg'),
        r'2,3,4,5-Tetraklorfenol': ('CP_2,3,4,5-TeCP', '2,3,4,5-Tetraklorfenol', 'mg/kg'),
        r'2,3,4,6-Tetraklorfenol': ('CP_2,3,4,6-TeCP', '2,3,4,6-Tetraklorfenol', 'mg/kg'),
        r'2,3,5,6-Tetraklorfenol': ('CP_2,3,5,6-TeCP', '2,3,5,6-Tetraklorfenol', 'mg/kg'),
        r'Pentaklorfenol': ('CP_PCP', 'Pentaklorfenol', 'mg/kg'),
        
        # CHLOROBENZENES
        r'Monoklorbensen': ('CB_MCB', 'Monoklorbensen', 'mg/kg'),
        r'1,2-Diklorbensen': ('CB_1,2-DCB', '1,2-Diklorbensen', 'mg/kg'),
        r'1,4-Diklorbensen': ('CB_1,4-DCB', '1,4-Diklorbensen', 'mg/kg'),
        r'1,2,3-Triklorbensen': ('CB_1,2,3-TCB', '1,2,3-Triklorbensen', 'mg/kg'),
        r'1,2,4-Triklorbensen': ('CB_1,2,4-TCB', '1,2,4-Triklorbensen', 'mg/kg'),
        r'1,3,5-Triklorbensen': ('CB_1,3,5-TCB', '1,3,5-Triklorbensen', 'mg/kg'),
        r'1,2,3,5\+1,2,4,5-Tetraklorbensen': ('CB_TeCB', '1,2,3,5+1,2,4,5-Tetraklorbensen', 'mg/kg'),
        r'Pentaklorbensen': ('CB_PeCB', 'Pentaklorbensen', 'mg/kg'),
        r'Heksaklorbensen': ('CB_HCB', 'Heksaklorbensen', 'mg/kg'),
        
        # CHLORINATED SOLVENTS
        r'Diklormetana?(?!\s*\()': ('VOC_DCM', 'Diklormetana', 'mg/kg'),
        r'Triklormetan\s*\(kloroform\)': ('VOC_Chloroform', 'Triklormetan (kloroform)', 'mg/kg'),
        r'Trikloreten': ('VOC_TCE', 'Trikloreten', 'mg/kg'),
        r'Tetraklormetana?': ('VOC_CCl4', 'Tetraklormetana', 'mg/kg'),
        r'Tetrakloreten': ('VOC_PCE', 'Tetrakloreten', 'mg/kg'),
        r'1,2-Dikloretan': ('VOC_1,2-DCA', '1,2-Dikloretan', 'mg/kg'),
        r'1,1,1-Trikloretan': ('VOC_1,1,1-TCA', '1,1,1-Trikloretan', 'mg/kg'),
        r'1,2-Dibrometan': ('VOC_EDB', '1,2-Dibrometan', 'mg/kg'),
        r'1,1,2-Trikloretan': ('VOC_1,1,2-TCA', '1,1,2-Trikloretan', 'mg/kg'),
        
        # PESTICIDES
        r'g-HCH\s*\(Lindan\)': ('PEST_Lindane', 'g-HCH (Lindan)', 'mg/kg'),
        r"o,p'-DDT": ('PEST_op-DDT', "o,p'-DDT", 'mg/kg'),
        r"p,p'-DDT": ('PEST_pp-DDT', "p,p'-DDT", 'mg/kg'),
        r"o,p'-DDD": ('PEST_op-DDD', "o,p'-DDD", 'mg/kg'),
        r"p,p'-DDD": ('PEST_pp-DDD', "p,p'-DDD", 'mg/kg'),
        r"o,p'-DDE": ('PEST_op-DDE', "o,p'-DDE", 'mg/kg'),
        r"p,p'-DDE": ('PEST_pp-DDE', "p,p'-DDE", 'mg/kg'),
    }
    
    # =========================================================================
    # LINE-BY-LINE PARSING
    # Match: ParamName <value> [uncertainty] unit
    # =========================================================================
    
    # Generic value pattern: captures value (with optional <) and optional uncertainty
    value_pattern = r'\s+([<n][\d.,n\.d]+|[\d.,]+)\s+(?:(\d+[.,]?\d*)\s+)?(mg/kgTS|mg/kg\s*TS|%)'
    
    for param_regex, (param_code, param_raw, expected_unit) in PARAM_MAP.items():
        if param_code in found_params:
            continue
            
        # Build full pattern: param name + value + unit
        full_pattern = param_regex + value_pattern
        match = re.search(full_pattern, clean_section, re.IGNORECASE)
        
        if match:
            value_str = match.group(1).replace(',', '.').strip()
            uncertainty_str = match.group(2) if match.group(2) else None
            
            result = build_result(sample_id, param_code, param_raw, value_str, 
                                  expected_unit, uncertainty_str)
            if result:
                results.append(result)
                found_params.add(param_code)
    
    return results


def build_result(sample_id: str, param_code: str, param_raw: str, 
                 value_str: str, unit: str, uncertainty_str: str = None) -> dict:
    """
    Build a result dictionary from parsed values.
    """
    below_limit = False
    loq = None
    uncertainty = None
    
    # Parse uncertainty
    if uncertainty_str:
        try:
            uncertainty = float(uncertainty_str.replace(',', '.'))
        except:
            pass
    
    # Parse value
    if value_str.startswith('<'):
        below_limit = True
        value_str = value_str[1:]
        try:
            loq = float(value_str)
            value = loq
        except:
            value = None
    elif value_str.lower() in ['n.d.', 'nd', 'n.d']:
        below_limit = True
        value = 0.0
    else:
        try:
            value = float(value_str)
        except:
            return None
    
    return {
        'sample_id': sample_id,
        'parameter': param_code,
        'parameter_raw': param_raw,
        'value': value,
        'unit': unit,
        'uncertainty': uncertainty,
        'below_limit': below_limit,
        'loq': loq,
    }


def create_wide_table(results_df: pd.DataFrame) -> pd.DataFrame:
    """
    Create a wide format table with:
    - Rows: parameters
    - Columns: sample IDs with unit in header (e.g., "p09-MOA-001 (mg/kg)")
    - Values: formatted as "value" or "<value" for below-limit
    
    Alerts if units vary within a sample column.
    """
    # Check for unit consistency within each sample
    unit_check = results_df.groupby('sample_id')['unit'].nunique()
    inconsistent_samples = unit_check[unit_check > 1]
    if len(inconsistent_samples) > 0:
        print(f"  WARNING: Mixed units found in samples: {list(inconsistent_samples.index)}")
        for sample_id in inconsistent_samples.index:
            units = results_df[results_df['sample_id'] == sample_id]['unit'].unique()
            print(f"    {sample_id}: {list(units)}")
    
    # Create formatted value strings (without units)
    def format_value(row):
        if pd.isna(row['value']):
            return ''
        
        # Format the numeric value
        val = row['value']
        if val == int(val):
            val_str = str(int(val))
        elif val < 0.01:
            val_str = f"{val:.4f}"
        elif val < 1:
            val_str = f"{val:.3f}"
        else:
            val_str = f"{val:.2f}".rstrip('0').rstrip('.')
        
        # Add < prefix for below-limit values
        if row['below_limit']:
            val_str = f"<{val_str}"
        
        return val_str
    
    results_df = results_df.copy()
    results_df['formatted'] = results_df.apply(format_value, axis=1)
    
    # Get predominant unit per sample (for header)
    sample_units = results_df.groupby('sample_id')['unit'].agg(
        lambda x: x.value_counts().index[0]  # Most common unit
    ).to_dict()
    
    # Pivot to wide format
    wide_df = results_df.pivot_table(
        index=['parameter', 'parameter_raw'],
        columns='sample_id',
        values='formatted',
        aggfunc='first'
    ).reset_index()
    
    # Rename columns for clarity
    wide_df.columns.name = None
    wide_df = wide_df.rename(columns={'parameter': 'param_code', 'parameter_raw': 'parameter'})
    
    # Rename sample columns to include unit in header
    sample_cols = [c for c in wide_df.columns if c.startswith('p09-')]
    sample_cols.sort()
    col_rename = {col: f"{col} ({sample_units.get(col, 'mg/kg')})" for col in sample_cols}
    wide_df = wide_df.rename(columns=col_rename)
    
    # Reorder columns
    new_sample_cols = [col_rename[c] for c in sample_cols]
    wide_df = wide_df[['param_code', 'parameter'] + new_sample_cols]
    
    return wide_df


def save_wide_table_xlsx(wide_df: pd.DataFrame, filepath: Path) -> None:
    """
    Save wide table as formatted Excel table for QA.
    - Formatted as Excel Table with filters
    - Alternating row colors
    - Column widths auto-adjusted
    - Header row frozen
    """
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    
    # Write dataframe to worksheet
    for r_idx, row in enumerate(dataframe_to_rows(wide_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Define table range
    max_row = len(wide_df) + 1  # +1 for header
    max_col = len(wide_df.columns)
    table_range = f"A1:{get_column_letter(max_col)}{max_row}"
    
    # Create table with style
    table = Table(displayName="ResultsTable", ref=table_range)
    style = TableStyleInfo(
        name="TableStyleMedium2",  # Blue alternating rows
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    
    # Auto-adjust column widths
    for col_idx, col in enumerate(wide_df.columns, 1):
        # Calculate max width based on content
        max_width = len(str(col))  # Header width
        for row_idx in range(2, max_row + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_width = max(max_width, len(str(cell_value)))
        
        # Set width with some padding, max 30 chars
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 2, 30)
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save
    wb.save(filepath)


def create_qa_workbook(filepath: Path, summary: dict, samples_df: pd.DataFrame,
                       results_df: pd.DataFrame, class_df: pd.DataFrame, 
                       dec_df: pd.DataFrame) -> None:
    """
    Create a QA workbook for manual verification of extracted data.
    
    Sheets:
    - Summary: Extraction metadata and file overview
    - QA Checklist: Items to verify with status columns
    - Samples QA: Sample data with QA columns
    - Classifications QA: Classification data with QA columns
    - Decisions QA: Decision data with QA columns
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # =========================================================================
    # SHEET 1: Summary
    # =========================================================================
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Title
    ws_summary['A1'] = f"QA Workbook: {summary['project']}"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:D1')
    
    # Extraction info
    summary_items = [
        ('Project Code', summary['project_code']),
        ('Extracted At', summary['extracted_at']),
        ('Source File', summary['source_file']),
        ('Notat Pages', summary['notat_pages']),
        ('Lab Report Pages', summary['lab_report_pages']),
        ('Lab', summary['lab']),
        ('Lab Reports', summary['lab_reports']),
        ('Sampling Dates', summary['sampling_dates']),
    ]
    
    for i, (label, value) in enumerate(summary_items, start=3):
        ws_summary[f'A{i}'] = label
        ws_summary[f'A{i}'].font = Font(bold=True)
        ws_summary[f'B{i}'] = value
    
    # Data counts
    ws_summary['A13'] = "Data Exports"
    ws_summary['A13'].font = Font(bold=True, size=12)
    
    export_items = [
        ('File', 'Rows', 'Description', 'Remarks', 'Date', 'Name'),
        ('samples.csv', summary['samples_count'], 'Sample metadata', '', '', ''),
        ('results.csv', summary['results_count'], 'Lab analysis results (long format)', '', '', ''),
        ('results_wide.csv', len(results_df['parameter'].unique()) if len(results_df) > 0 else 0, 'Lab results (wide format for review)', '', '', ''),
        ('results_wide.xlsx', len(results_df['parameter'].unique()) if len(results_df) > 0 else 0, 'Lab results (Excel table)', '', '', ''),
        ('classifications.csv', summary['classifications_count'], 'Tilstandsklasse per sample', '', '', ''),
        ('decisions.csv', summary['decisions_count'], 'Handling decisions per sample', '', '', ''),
        ('extraction_summary.csv', 1, 'Extraction metadata', '', '', ''),
    ]
    
    for i, row_data in enumerate(export_items, start=14):
        for j, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=i, column=j, value=value)
            if i == 14:  # Header row
                cell.font = header_font
                cell.fill = header_fill
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 10
    ws_summary.column_dimensions['C'].width = 40
    ws_summary.column_dimensions['D'].width = 30
    ws_summary.column_dimensions['E'].width = 12
    ws_summary.column_dimensions['F'].width = 15
    
    # =========================================================================
    # SHEET 2: QA Checklist
    # =========================================================================
    ws_qa = wb.create_sheet("QA Checklist")
    
    checklist_items = [
        ('Category', 'Check Item', 'Status', 'Verified By', 'Date', 'Notes'),
        ('Samples', 'All samples from source document captured', '', '', '', ''),
        ('Samples', 'Sample IDs match lab report numbers', '', '', '', ''),
        ('Samples', 'Sample dates correct', '', '', '', ''),
        ('Samples', 'Profile locations match source', '', '', '', ''),
        ('Results', 'All parameters extracted for each sample', '', '', '', ''),
        ('Results', 'Values match source PDF', '', '', '', ''),
        ('Results', 'Below-limit values (<) correctly flagged', '', '', '', ''),
        ('Results', 'Units correct (mg/kg, %)', '', '', '', ''),
        ('Results', 'Uncertainties captured where available', '', '', '', ''),
        ('Classifications', 'Tilstandsklasse matches source tables', '', '', '', ''),
        ('Classifications', 'Limiting parameters identified correctly', '', '', '', ''),
        ('Classifications', 'Classification basis documented', '', '', '', ''),
        ('Decisions', 'Handling decisions match notat text', '', '', '', ''),
        ('Decisions', 'Destinations correct (deponi, gjenbruk, etc)', '', '', '', ''),
        ('Decisions', 'Responsible parties assigned correctly', '', '', '', ''),
        ('General', 'No duplicate samples', '', '', '', ''),
        ('General', 'No missing critical data', '', '', '', ''),
        ('General', 'Data ready for aggregation', '', '', '', ''),
    ]
    
    for i, row in enumerate(checklist_items, start=1):
        for j, value in enumerate(row, start=1):
            cell = ws_qa.cell(row=i, column=j, value=value)
            cell.border = thin_border
            if i == 1:  # Header
                cell.font = header_font
                cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Set column widths
    ws_qa.column_dimensions['A'].width = 15
    ws_qa.column_dimensions['B'].width = 45
    ws_qa.column_dimensions['C'].width = 12
    ws_qa.column_dimensions['D'].width = 15
    ws_qa.column_dimensions['E'].width = 12
    ws_qa.column_dimensions['F'].width = 40
    
    # Freeze header
    ws_qa.freeze_panes = 'A2'
    
    # =========================================================================
    # SHEET 3: Samples QA
    # =========================================================================
    ws_samples = wb.create_sheet("Samples QA")
    
    # Add QA columns to samples
    samples_qa = samples_df.copy()
    samples_qa['QA_Status'] = ''
    samples_qa['QA_Notes'] = ''
    
    for i, row in enumerate(dataframe_to_rows(samples_qa, index=False, header=True), start=1):
        for j, value in enumerate(row, start=1):
            cell = ws_samples.cell(row=i, column=j, value=value)
            if i == 1:
                cell.font = header_font
                cell.fill = header_fill
    
    ws_samples.freeze_panes = 'A2'
    
    # Auto-width columns
    for col_idx, col in enumerate(samples_qa.columns, start=1):
        ws_samples.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col)) + 2, 12)
    
    # =========================================================================
    # SHEET 4: Classifications QA
    # =========================================================================
    ws_class = wb.create_sheet("Classifications QA")
    
    class_qa = class_df.copy()
    class_qa['QA_Status'] = ''
    class_qa['QA_Notes'] = ''
    
    for i, row in enumerate(dataframe_to_rows(class_qa, index=False, header=True), start=1):
        for j, value in enumerate(row, start=1):
            cell = ws_class.cell(row=i, column=j, value=value)
            if i == 1:
                cell.font = header_font
                cell.fill = header_fill
    
    ws_class.freeze_panes = 'A2'
    
    for col_idx, col in enumerate(class_qa.columns, start=1):
        ws_class.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col)) + 2, 15)
    
    # =========================================================================
    # SHEET 5: Decisions QA
    # =========================================================================
    ws_dec = wb.create_sheet("Decisions QA")
    
    dec_qa = dec_df.copy()
    dec_qa['QA_Status'] = ''
    dec_qa['QA_Notes'] = ''
    
    for i, row in enumerate(dataframe_to_rows(dec_qa, index=False, header=True), start=1):
        for j, value in enumerate(row, start=1):
            cell = ws_dec.cell(row=i, column=j, value=value)
            if i == 1:
                cell.font = header_font
                cell.fill = header_fill
    
    ws_dec.freeze_panes = 'A2'
    
    for col_idx, col in enumerate(dec_qa.columns, start=1):
        ws_dec.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col)) + 2, 15)
    
    # Save workbook
    wb.save(filepath)


def extract():
    """Main extraction function for Project 09."""
    print(f"{'='*60}")
    print(f"Extracting: {PROJECT_NAME}")
    print(f"{'='*60}")
    
    # Ensure output directory exists
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    pdf_path = INBOX_DIR / PDF_FILE
    
    if not pdf_path.exists():
        print(f"ERROR: PDF file not found: {pdf_path}")
        return False
    
    print(f"\nSource: {PDF_FILE}")
    
    # ============================================================
    # EXTRACT RAW TEXT
    # ============================================================
    print("\n[1/6] Extracting raw text...")
    
    try:
        full_text = extract_text(pdf_path)
        text_output = OUTPUT_DIR / 'p09_raw_text.txt'
        text_output.write_text(full_text, encoding='utf-8')
        print(f"  Saved: p09_raw_text.txt")
    except Exception as e:
        print(f"ERROR: {e}")
        return False
    
    # ============================================================
    # EXTRACT NOTAT PAGES (2-5)
    # ============================================================
    print(f"\n[2/6] Extracting notat (pages {NOTAT_PAGES[0]}-{NOTAT_PAGES[-1]})...")
    
    try:
        notat_text = extract_pages(pdf_path, NOTAT_PAGES)
        notat_output = OUTPUT_DIR / 'p09_notat_pages_2-5.txt'
        notat_output.write_text(notat_text, encoding='utf-8')
        print(f"  Saved: p09_notat_pages_2-5.txt")
    except Exception as e:
        print(f"  Warning: {e}")
    
    # ============================================================
    # EXTRACT LAB REPORT PAGES (6-30)
    # ============================================================
    print(f"\n[3/6] Extracting lab reports (pages {LAB_REPORT_PAGES[0]}-{LAB_REPORT_PAGES[-1]})...")
    
    try:
        lab_text = extract_pages(pdf_path, LAB_REPORT_PAGES)
        lab_output = OUTPUT_DIR / 'p09_lab_reports_pages_6-30.txt'
        lab_output.write_text(lab_text, encoding='utf-8')
        print(f"  Saved: p09_lab_reports_pages_6-30.txt")
    except Exception as e:
        print(f"  Warning: {e}")
        lab_text = full_text  # Fallback to full text
    
    # ============================================================
    # BUILD SAMPLES TABLE
    # ============================================================
    print("\n[4/6] Building samples.csv...")
    
    samples_data = []
    for s in SAMPLES:
        samples_data.append({
            'sample_id': s['sample_id'],
            'project_code': PROJECT_CODE,
            'sample_date': s['sample_date'],
            'location_type': s['location_type'],
            'profile_start': s['profile_start'] if s['profile_start'] else '',
            'profile_end': s['profile_end'] if s['profile_end'] else '',
            'tunnel_name': TUNNEL_NAME,
            'sample_type': s['sample_type'],
            'lab_reference': s['lab_reference'],
            'sampler': 'SVV Region Sør',
        })
    
    samples_df = pd.DataFrame(samples_data)
    save_to_csv(samples_df, OUTPUT_DIR / 'p09_samples.csv')
    print(f"  Saved: p09_samples.csv ({len(samples_df)} samples)")
    
    # ============================================================
    # PARSE AND BUILD RESULTS TABLE
    # ============================================================
    print("\n[5/6] Parsing lab results and building results.csv...")
    
    results = parse_lab_results(lab_text)
    print(f"  Parsed {len(results)} result entries from lab reports")
    
    if results:
        results_df = pd.DataFrame(results)
        results_df = results_df[['sample_id', 'parameter', 'parameter_raw', 'value', 
                                  'unit', 'uncertainty', 'below_limit', 'loq']]
        save_to_csv(results_df, OUTPUT_DIR / 'p09_results.csv')
        print(f"  Saved: p09_results.csv ({len(results_df)} results)")
        
        # Summary by sample
        sample_counts = results_df.groupby('sample_id').size()
        for sid, count in sample_counts.items():
            print(f"    {sid}: {count} parameters")
        
        # Create wide format table (samples as columns, parameters as rows)
        wide_df = create_wide_table(results_df)
        save_to_csv(wide_df, OUTPUT_DIR / 'p09_results_wide.csv')
        print(f"  Saved: p09_results_wide.csv ({len(wide_df)} parameters x {len(wide_df.columns)-2} samples)")
        
        # Export as formatted Excel table for QA
        save_wide_table_xlsx(wide_df, OUTPUT_DIR / 'p09_results_wide.xlsx')
        print(f"  Saved: p09_results_wide.xlsx (formatted table for QA)")
    else:
        print("  WARNING: No results parsed!")
        results_df = pd.DataFrame()
    
    # ============================================================
    # BUILD CLASSIFICATIONS TABLE
    # ============================================================
    print("\n[6/6] Building classifications.csv and decisions.csv...")
    
    class_data = []
    for c in CLASSIFICATIONS:
        sample_id = SAMPLE_KEY_TO_ID.get(c['sample_key'], f"MOA-{c['sample_key']}")
        # Join lists with semicolon for CSV compatibility
        params = c['limiting_parameters']
        class_data.append({
            'sample_id': sample_id,
            'tilstandsklasse': c['tilstandsklasse'],
            'limiting_parameters': ';'.join(params) if params else '',
            'classification_basis': 'TA-2553/2009',
        })
    
    class_df = pd.DataFrame(class_data)
    save_to_csv(class_df, OUTPUT_DIR / 'p09_classifications.csv')
    print(f"  Saved: p09_classifications.csv ({len(class_df)} classifications)")
    
    # ============================================================
    # BUILD DECISIONS TABLE
    # ============================================================
    dec_data = []
    for d in DECISIONS:
        sample_id = SAMPLE_KEY_TO_ID.get(d['sample_key'], f"MOA-{d['sample_key']}")
        dec_data.append({
            'sample_id': sample_id,
            'decision': d['decision'],
            'destination': d['destination'],
            'notes': d['notes'],
        })
    
    dec_df = pd.DataFrame(dec_data)
    save_to_csv(dec_df, OUTPUT_DIR / 'p09_decisions.csv')
    print(f"  Saved: p09_decisions.csv ({len(dec_df)} decisions)")
    
    # ============================================================
    # GENERATE EXTRACTION SUMMARY
    # ============================================================
    summary = {
        'project': PROJECT_NAME,
        'project_code': PROJECT_CODE,
        'extracted_at': datetime.now().isoformat(),
        'source_file': PDF_FILE,
        'notat_pages': f"{NOTAT_PAGES[0]}-{NOTAT_PAGES[-1]}",
        'lab_report_pages': f"{LAB_REPORT_PAGES[0]}-{LAB_REPORT_PAGES[-1]}",
        'samples_count': len(samples_df),
        'results_count': len(results_df) if len(results) > 0 else 0,
        'classifications_count': len(class_df),
        'decisions_count': len(dec_df),
        'sampling_dates': '2017-09-20, 2017-09-28',
        'lab': 'ALS Laboratory Group Norway AS',
        'lab_reports': 'N1715906, N1716546',
    }
    
    summary_df = pd.DataFrame([summary])
    save_to_csv(summary_df, OUTPUT_DIR / 'p09_extraction_summary.csv')
    
    # ============================================================
    # GENERATE QA WORKBOOK (only if not exists to avoid overwriting notes)
    # ============================================================
    qa_workbook_path = OUTPUT_DIR / 'p09_QA_workbook.xlsx'
    if qa_workbook_path.exists():
        print(f"  SKIPPED: p09_QA_workbook.xlsx already exists (preserving manual notes)")
    else:
        create_qa_workbook(
            qa_workbook_path,
            summary,
            samples_df,
            results_df if len(results) > 0 else pd.DataFrame(),
            class_df,
            dec_df
        )
        print(f"  Saved: p09_QA_workbook.xlsx (for manual QA documentation)")
    
    # ============================================================
    # FINAL REPORT
    # ============================================================
    print(f"\n{'='*60}")
    print("EXTRACTION COMPLETE")
    print(f"{'='*60}")
    print(f"\nOutput files in: {OUTPUT_DIR}")
    print(f"  - p09_samples.csv          ({summary['samples_count']} rows)")
    print(f"  - p09_results.csv          ({summary['results_count']} rows)")
    print(f"  - p09_classifications.csv  ({summary['classifications_count']} rows)")
    print(f"  - p09_decisions.csv        ({summary['decisions_count']} rows)")
    print(f"  - p09_extraction_summary.csv")
    print(f"  - p09_raw_text.txt")
    print(f"  - p09_notat_pages_2-5.txt")
    print(f"  - p09_lab_reports_pages_6-30.txt")
    print(f"{'='*60}")
    
    return True


if __name__ == '__main__':
    success = extract()
    sys.exit(0 if success else 1)
