"""
Project 01 - Høvik Tunnel (E18 Vestkorridoren E102) Extraction Script

Extracts data from Excel file containing ALS lab results.
Source: SVV E18 Vestkorridoren E102 - Høvik tunnel bunnrensk.xlsx

Tunnels:
- HTØ: Høvik Øst (East)
- HTV: Høvik Tverrslag (Cross passage)
- HV: Høvik Vest (West)

Output:
- p01_samples.csv
- p01_results.csv
- p01_classifications.csv
- p01_decisions.csv
- p01_results_wide.csv
- p01_results_wide.xlsx
"""

import sys
from pathlib import Path
import pandas as pd
import openpyxl

# Add parent to path for lib imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from lib.export import save_to_csv
from lib.chemistry import normalize_parameter_name, PARAMETER_ALIASES
from lib.excel_utils import create_wide_table, save_wide_table_xlsx
from lib.qa import create_qa_workbook
from lib.dataframes import (
    create_samples_df, create_results_df,
    create_classifications_df, create_decisions_df
)

# ============================================================
# PROJECT CONFIGURATION
# ============================================================

PROJECT_CODE = '01_e18-e102'
PROJECT_NAME = 'E18 Vestkorridoren E102 - Høvik'
TUNNEL_NAME = 'Høviktunnelen'
PROJECT_PREFIX = 'p01-'

BASE_DIR = Path(__file__).parent.parent.parent
INBOX_DIR = BASE_DIR / '00_inbox' / 'SVV' / 'E18 Vestkorridoren E102'
OUTPUT_DIR = BASE_DIR / '01_projects' / '01_e18-e102' / 'extracted'

EXCEL_FILE = 'Høvik tunnel bunnrensk.xlsx'
KORNFORDELING_FILE = 'Kornfordeling Høvik bunnrensk mai 2025.xlsx'

# Row indices (0-based in openpyxl: row 1 = index 1)
ROW_TUNNEL = 2      # Tunnel name (Høvik Øst, etc.)
ROW_ZONE = 3        # Zone (Sone 1, Sone 2)
ROW_SAMPLE = 4      # Sample name (prøvenavn)
ROW_REP_FOR = 5     # Representative for (pel range)
ROW_ORDER_NR = 6    # ALS order number
ROW_DATE = 7        # Date of sampling
ROW_HEADER = 8      # Header row (Analyse, Enhet, Grenseverdi...)
ROW_DATA_START = 9  # First data row

# Columns to skip (not chemical parameters)
SKIP_PARAMS = {'Tørrstoff', 'TOC'}

# Color codes for decoding tilstandsklasse from source Excel cell backgrounds
INPUT_COLOR_TO_TK = {
    'FF00B0F0': 1,   # Light blue - TK1
    'FF92D050': 2,   # Green - TK2
    'FFFFFF00': 3,   # Yellow - TK3
    'FFFFC000': 4,   # Orange - TK4
    'FFFF0000': 5,   # Red - TK5
}

# Hardcoded grain size distribution for HTV bunnrensk (May 2025)
# Format: (sieve_size_mm, cumulative_percent_passing)
KORNFORDELING_DATA = [
    (0, 0),
    (0.002, 0.22),
    (0.004, 0.61),
    (0.008, 1.11),
    (0.016, 1.59),
    (0.032, 2.03),
    (0.063, 2.4),
    (0.125, 2.91),
    (0.25, 3.51),
    (0.5, 4.23),
    (1, 5.41),
    (2, 6.79),
    (4, 8.65),
    (8, 12.72),
    (16, 19.59),
    (31.5, 28.96),
    (63, 72.04),
    (100, 100),  # >63mm fraction
]


def parse_value(val):
    """
    Parse a value from the Excel file.
    Returns (numeric_value, below_limit, loq)
    """
    if pd.isna(val) or val == '' or val == 'nd':
        return None, False, None
    
    val_str = str(val).strip()
    
    # Handle below-limit values like "< 0,20" or "<0.20" or "<2.0"
    if val_str.startswith('<'):
        val_str = val_str[1:].strip()
        val_str = val_str.replace(',', '.').replace(' ', '')
        try:
            loq = float(val_str)
            return loq, True, loq
        except:
            return None, True, None
    
    # Regular numeric value
    val_str = val_str.replace(',', '.')
    try:
        return float(val_str), False, None
    except:
        return None, False, None


def get_tilstandsklasse_from_color(color_rgb: str) -> int:
    """Get tilstandsklasse from cell color"""
    if color_rgb and color_rgb in INPUT_COLOR_TO_TK:
        return INPUT_COLOR_TO_TK[color_rgb]
    return None


def parse_profile_from_rep_for(rep_for: str) -> tuple:
    """Extract profile start/end from 'Representative for' field like 'P8460 - P8520'"""
    import re
    if not rep_for:
        return None, None
    # Look for patterns like P8460 - P8520 or P8460-P8520
    numbers = re.findall(r'P?(\d{4,5})', str(rep_for))
    if len(numbers) >= 2:
        return float(numbers[0]), float(numbers[1])
    elif len(numbers) == 1:
        return float(numbers[0]), None
    return None, None


def extract_data(excel_path: Path) -> tuple[list, list]:
    """
    Extract samples and results from the Høvik tunnel Excel file.
    
    Returns:
        Tuple of (samples_list, results_list)
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    samples = []
    results = []
    
    # Find all sample columns (columns with data in the sample name row)
    sample_cols = []
    for col in range(4, ws.max_column + 1):  # Start from col 4 (D)
        sample_cell = ws.cell(row=ROW_SAMPLE, column=col)
        if sample_cell.value and str(sample_cell.value).strip():
            sample_name = str(sample_cell.value).strip()
            # Skip empty or header columns
            if sample_name and sample_name != 'Prøvenavn':
                sample_cols.append((col, sample_name))
    
    print(f"Found {len(sample_cols)} samples")
    
    # Extract sample info
    for col, sample_name in sample_cols:
        # Get metadata
        tunnel = ws.cell(row=ROW_TUNNEL, column=col).value or ''
        zone = ws.cell(row=ROW_ZONE, column=col).value or ''
        rep_for = ws.cell(row=ROW_REP_FOR, column=col).value or ''
        order_nr = ws.cell(row=ROW_ORDER_NR, column=col).value or ''
        sample_date = ws.cell(row=ROW_DATE, column=col).value
        
        # Parse tunnel code and location_type from sample name
        # location_type: grøft if "grøft" in name, pumpesump, otherwise vegbane
        if 'grøft' in sample_name.lower():
            location_type = 'grøft'
        elif 'Pumpesump' in sample_name:
            location_type = 'pumpesump'
        else:
            location_type = 'vegbane'
        
        # Create sample ID
        clean_name = sample_name.replace(' ', '').replace('#', '-')
        sample_id = f"{PROJECT_PREFIX}{clean_name}"
        
        # Convert date
        if isinstance(sample_date, pd.Timestamp):
            sample_date_str = sample_date.strftime('%Y-%m-%d')
        elif hasattr(sample_date, 'strftime'):
            sample_date_str = sample_date.strftime('%Y-%m-%d')
        else:
            sample_date_str = ''
        
        # Parse profile from representative_for field
        profile_start, profile_end = parse_profile_from_rep_for(rep_for)
        
        samples.append({
            'sample_id': sample_id,
            'project_code': PROJECT_CODE,
            'sample_date': sample_date_str,
            'location_type': location_type,
            'profile_start': profile_start,
            'profile_end': profile_end,
            'tunnel_name': TUNNEL_NAME,
            'sample_type': 'bunnrensk',
            'lab_reference': str(order_nr).strip() if order_nr else '',
            'sampler': 'Skanska',
            'remark': '',
        })
        
        # Extract results for this sample
        for row in range(ROW_DATA_START, ws.max_row + 1):
            param_cell = ws.cell(row=row, column=1)
            param_name = param_cell.value
            
            if not param_name or pd.isna(param_name):
                continue
            
            param_name = str(param_name).strip()
            if param_name in SKIP_PARAMS:
                continue
            
            # Get unit
            unit_cell = ws.cell(row=row, column=2)
            unit = str(unit_cell.value).strip() if unit_cell.value else 'mg/kg'
            
            # Get value and color
            value_cell = ws.cell(row=row, column=col)
            value, below_limit, loq = parse_value(value_cell.value)
            
            if value is None:
                continue
            
            # Get tilstandsklasse from color
            fill = value_cell.fill
            color_rgb = None
            if fill and fill.fgColor and fill.fgColor.rgb:
                color_rgb = fill.fgColor.rgb
            tilstandsklasse = get_tilstandsklasse_from_color(color_rgb)
            
            # Get parameter code (uses shared PARAMETER_ALIASES via normalize_parameter_name)
            param_code = normalize_parameter_name(param_name)
            
            results.append({
                'sample_id': sample_id,
                'parameter': param_code,
                'parameter_raw': param_name,
                'value': value,
                'unit': unit,
                'uncertainty': None,
                'below_limit': below_limit,
                'loq': loq,
                'analysis_type': 'totalanalyse',
                'tilstandsklasse': tilstandsklasse,  # Stored for classification calculation
            })
    
    return samples, results


def extract_kornfordeling() -> tuple[dict, list]:
    """
    Return grain size distribution data for HTV bunnrensk (May 2025).
    Uses KORNFORDELING_DATA from config section.
    
    Returns:
        Tuple of (sample_dict, results_list)
    """
    sample_id = f"{PROJECT_PREFIX}BR-HTV-kornfordeling"
    sample = {
        'sample_id': sample_id,
        'project_code': PROJECT_CODE,
        'sample_date': '2025-05-16',
        'location_type': 'vegbane',
        'profile_start': None,
        'profile_end': None,
        'tunnel_name': TUNNEL_NAME,
        'sample_type': 'bunnrensk',
        'lab_reference': 'NO25115',
        'sampler': 'Skanska',
        'remark': '',
    }
    
    results = []
    for size_mm, cumul_pct in KORNFORDELING_DATA:
        param_name = f"Kornfordeling_{size_mm}mm"
        results.append({
            'sample_id': sample_id,
            'parameter': param_name,
            'parameter_raw': f'Kumulativ % passerende {size_mm} mm',
            'value': cumul_pct,
            'unit': '%',
            'uncertainty': None,
            'below_limit': False,
            'loq': None,
            'analysis_type': 'kornfordeling',
        })
    
    return sample, results


def main():
    """Main extraction function"""
    print(f"=" * 60)
    print(f"Project: {PROJECT_NAME}")
    print(f"=" * 60)
    
    excel_path = INBOX_DIR / EXCEL_FILE
    print(f"\nSource file: {excel_path}")
    
    if not excel_path.exists():
        print(f"ERROR: Excel file not found: {excel_path}")
        return
    
    # Create output directory
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    # Extract data
    print("\nExtracting data from Excel...")
    samples_list, results_list = extract_data(excel_path)
    
    # Extract kornfordeling (hardcoded values)
    print("\nAdding kornfordeling...")
    korn_sample, korn_results = extract_kornfordeling()
    samples_list.append(korn_sample)
    results_list.extend(korn_results)
    print(f"  - Added {len(korn_results)} grain size fractions")
    
    # Create DataFrames with standard schema
    samples_df = create_samples_df(samples_list)
    
    # Results need tilstandsklasse temporarily for classification
    results_df_full = pd.DataFrame(results_list)
    results_df = create_results_df(results_list)
    
    print(f"\nExtracted:")
    print(f"  - {len(samples_df)} samples")
    print(f"  - {len(results_df)} results")
    
    # Show sample breakdown by location
    if len(samples_df) > 0:
        print(f"\nSamples by location:")
        for loc, count in samples_df['location_type'].value_counts().items():
            print(f"  - {loc}: {count}")
    
    # Create classifications and decisions using schema functions
    classifications_df = create_classifications_df(samples_df, results_df_full)
    decisions_df = create_decisions_df(samples_df, results_df_full)
    decisions_df['notes'] = 'Bestemmes på et senere tidspunkt'
    
    # Show classification summary
    if len(classifications_df) > 0:
        print(f"\nSample tilstandsklasse (based on highest per sample):")
        tk_counts = classifications_df['tilstandsklasse'].value_counts().sort_index()
        for tk, count in tk_counts.items():
            if pd.notna(tk):
                print(f"  - TK{int(tk)}: {count} samples")
    
    if len(decisions_df) > 0:
        print(f"\nDecision summary:")
        decision_counts = decisions_df['decision'].value_counts()
        for decision, count in decision_counts.items():
            print(f"  - {decision}: {count} samples")
    
    # Create wide table
    print("\nCreating wide table...")
    wide_df = create_wide_table(results_df, PROJECT_PREFIX)
    print(f"  - {len(wide_df)} parameters in wide table")
    
    # Save outputs
    print("\nSaving outputs...")
    from datetime import datetime
    
    save_to_csv(samples_df, OUTPUT_DIR / 'p01_samples.csv')
    save_to_csv(results_df, OUTPUT_DIR / 'p01_results.csv')
    save_to_csv(classifications_df, OUTPUT_DIR / 'p01_classifications.csv')
    save_to_csv(decisions_df, OUTPUT_DIR / 'p01_decisions.csv')
    save_to_csv(wide_df, OUTPUT_DIR / 'p01_results_wide.csv')
    
    # Also save wide table as Excel
    save_wide_table_xlsx(wide_df, OUTPUT_DIR / 'p01_results_wide.xlsx')
    
    # Create extraction summary
    summary = {
        'project': PROJECT_NAME,
        'project_code': PROJECT_CODE,
        'extracted_at': datetime.now().isoformat(),
        'source_file': EXCEL_FILE,
        'samples_count': len(samples_df),
        'results_count': len(results_df),
        'classifications_count': len(classifications_df),
        'decisions_count': len(decisions_df),
        'lab': 'ALS',
    }
    
    # QA Workbook (with timestamp suffix)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    create_qa_workbook(
        OUTPUT_DIR / f'p01_QA_workbook_{timestamp}.xlsx',
        summary,
        samples_df,
        results_df,
        classifications_df,
        decisions_df
    )
    print(f"  Saved: p01_QA_workbook_{timestamp}.xlsx")
    
    print(f"\nDone! Output saved to {OUTPUT_DIR}")


if __name__ == '__main__':
    main()
