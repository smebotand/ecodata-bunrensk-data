"""
Project 18 - Hestnestunnelen Extraction Script

Extracts data from Excel file containing Eurofins lab results.
Source: BaneNOR Hestnestunnelen - Analyseresultater bunnrenskprøver.xlsx

Sheets:
- Hovedløp (main tunnel)
- Tverrslag Nord (north cross passage)  
- Tverrslag Nord Rømning (north escape tunnel)
- Tverrslag Sør (south cross passage)

Output:
- p18_samples.csv
- p18_results.csv
- p18_results_wide.csv
- p18_results_wide.xlsx
- p18_classifications.csv
- p18_decisions.csv
- p18_extraction_summary.csv
"""

import sys
import re
from pathlib import Path
import pandas as pd
from datetime import datetime

# Add parent to path for lib imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from lib.export import save_to_csv
from lib.chemistry import EUROFINS_ALI_COLUMN_MAP, EUROFINS_LEACHING_PARAM_MAP, get_parameter_code

# ============================================================
# PROJECT CONFIGURATION
# ============================================================

PROJECT_CODE = '18_hestnestunnelen'
PROJECT_NAME = 'Hestnestunnelen'
TUNNEL_NAME = 'Hestnestunnelen'

BASE_DIR = Path(__file__).parent.parent.parent
INBOX_DIR = BASE_DIR / '00_inbox' / 'BaneNOR' / 'Hestnestunnelen'
OUTPUT_DIR = BASE_DIR / '01_projects' / '18_hestnestunnelen' / 'extracted'

EXCEL_FILE = 'Hestnestunnelen - Analyseresultater bunnrenskprøver.xlsx'

# PDF file with leaching test results (ristetest and kolonnetest)
PDF_FILE = 'VD38199-VD-NO-0019-Gjenbruk av kjemisk rene bunnrenskmasser fra KS-1 Hestnestunnelen (1).pdf'
PDF_LEACHING_PAGES = list(range(34, 38))  # Pages 34-37 (0-indexed: 33-36)

# Sheet names and their location codes
SHEETS = {
    'Hovedløp': 'HL',           # Main tunnel
    'Tverrslag Nord': 'TVN',    # North cross passage
    'Tverrslag Nord Rømning': 'RØ',  # Escape tunnel
    'Tverrslag Sør': 'TVS',     # South cross passage
}

# Columns to skip (not chemical parameters)
SKIP_COLUMNS = {
    'Eurofins oppdragsmerking', 'Eurofins prøvenummer', 'Prøvemerking',
    'Fra pel', 'Til Pel', 'Test/utgått', 'Kommentar', 'Merknad'
}

# Color codes for tilstandsklasse (from Prøvemerking cell background)
# Based on standard Norwegian environmental color coding
COLOR_TO_TILSTANDSKLASSE = {
    'FF00B0F0': 1,   # Light blue
    'FF92D050': 2,   # Green
    'FFFFFF00': 3,   # Yellow
    'FFFFC000': 4,   # Orange
    'FFFF0000': 5,   # Red
    'FFEF3A3F': 5,   # Red variant
}


def parse_value(val):
    """
    Parse a value from the Excel file.
    Returns (numeric_value, below_limit, loq)
    """
    if pd.isna(val) or val == '' or val == 'nd' or val == 'Utgår':
        return None, False, None
    
    val_str = str(val).strip()
    
    # Handle below-limit values like "< 0,20" or "<0.20"
    if val_str.startswith('<'):
        val_str = val_str[1:].strip()
        val_str = val_str.replace(',', '.').replace(' ', '')
        try:
            loq = float(val_str)
            return loq, True, loq
        except:
            return None, True, None
    
    # Regular numeric value
    val_str = val_str.replace(',', '.')
    try:
        return float(val_str), False, None
    except:
        return None, False, None


def parse_leaching_test_pdf(pdf_path: Path, pages: list) -> list:
    """
    Parse leaching test results (ristetest L/S=10 and kolonnetest L/S=0,1) from Eurofins PDF.
    
    Args:
        pdf_path: Path to the PDF file
        pages: List of page numbers (1-indexed) to extract
        
    Returns:
        List of result dictionaries
    """
    import pdfplumber
    
    results = []
    
    with pdfplumber.open(pdf_path) as pdf:
        # Extract text from specified pages
        text = ''
        for page_num in pages:
            if page_num <= len(pdf.pages):
                page_text = pdf.pages[page_num - 1].extract_text() or ''
                text += page_text + '\n'
        
        # Find sample ID from Prøvemerking
        sample_match = re.search(r'Prøvemerking:\s*(\d+-\d+)', text)
        if not sample_match:
            print(f"  WARNING: Could not find sample ID in PDF pages {pages}")
            return results
        
        provemerking = sample_match.group(1)
        sample_id = f"p18-HL-{provemerking}"
        print(f"  Found leaching test sample: {provemerking}")
        
        # Parse ristetest results (L/S=10) - units are mg/kg TS
        # Format: "a) Parameter L/S=10 <value mg/kg TS LOQ [MU]"
        ristetest_pattern = r'a\)[\*]?\s+([A-Za-zæøåÆØÅ\s\(\)]+)\s+L/S=10\s+(<?\d+[.,]?\d*)\s+(mg/kg\s*TS|mS/m)\s+(\d+[.,]?\d*)?'
        
        for match in re.finditer(ristetest_pattern, text):
            param_raw = match.group(1).strip()
            value_str = match.group(2).replace(',', '.')
            unit = match.group(3).replace(' ', '').replace('TS', '')
            loq_str = match.group(4)
            
            # Skip non-parameter entries
            if param_raw in ['pH', 'Konduktivitet']:
                param_code = EUROFINS_LEACHING_PARAM_MAP.get(param_raw, param_raw)
            else:
                param_code = EUROFINS_LEACHING_PARAM_MAP.get(param_raw)
                if not param_code:
                    continue
            
            below_limit = value_str.startswith('<')
            if below_limit:
                value_str = value_str[1:]
            
            try:
                value = float(value_str)
            except:
                continue
            
            loq = None
            if loq_str:
                try:
                    loq = float(loq_str.replace(',', '.'))
                except:
                    pass
            if below_limit and loq is None:
                loq = value
            
            results.append({
                'sample_id': sample_id,
                'parameter': param_code,
                'parameter_raw': f"{param_raw} L/S=10",
                'value': value,
                'unit': unit,
                'uncertainty': None,
                'below_limit': below_limit,
                'loq': loq,
                'analysis_type': 'ristetest',
            })
        
        # Parse kolonnetest results (L/S=0,1) - units are mg/l
        # These appear after the ristetest section, without L/S in the line
        # Format: "a)* Parameter value mg/l LOQ [MU]"
        kolonnetest_section = text.split('L/S=0,1')
        if len(kolonnetest_section) > 1:
            kolon_text = kolonnetest_section[-1]  # Take text after last L/S=0,1 marker
            
            # Pattern for kolonnetest results (mg/l unit indicates kolonnetest)
            kolon_pattern = r'a\)[\*]?\s+([A-Za-zæøåÆØÅ\s\(\)]+?)\s+(<?\d+[.,]?\d*)\s+(mg/l|mS/m)\s+(\d+[.,]?\d*)?'
            
            for match in re.finditer(kolon_pattern, text):
                param_raw = match.group(1).strip()
                value_str = match.group(2).replace(',', '.')
                unit = match.group(3)
                loq_str = match.group(4)
                
                # Only process if unit is mg/l (kolonnetest)
                if unit != 'mg/l':
                    continue
                
                param_code = EUROFINS_LEACHING_PARAM_MAP.get(param_raw)
                if not param_code:
                    continue
                
                below_limit = value_str.startswith('<')
                if below_limit:
                    value_str = value_str[1:]
                
                try:
                    value = float(value_str)
                except:
                    continue
                
                loq = None
                if loq_str:
                    try:
                        loq = float(loq_str.replace(',', '.'))
                    except:
                        pass
                if below_limit and loq is None:
                    loq = value
                
                results.append({
                    'sample_id': sample_id,
                    'parameter': param_code,
                    'parameter_raw': f"{param_raw} L/S=0,1",
                    'value': value,
                    'unit': unit,
                    'uncertainty': None,
                    'below_limit': below_limit,
                    'loq': loq,
                    'analysis_type': 'kolonnetest',
                })
    
    return results


def read_sheet(filepath: Path, sheet_name: str, location_code: str, ws_openpyxl=None) -> tuple:
    """
    Read a single sheet from the Excel file.
    Returns (samples_list, results_list, classifications_list)
    
    Args:
        filepath: Path to Excel file
        sheet_name: Name of the sheet
        location_code: Short code for the location (e.g., 'HL', 'TVN')
        ws_openpyxl: Optional openpyxl worksheet for reading cell colors
    """
    # Read units from row 2 (0-indexed) - this row contains units like "mg/kg TS", "%"
    # Row 3 has parameter names, row 2 has units
    df_header = pd.read_excel(filepath, sheet_name=sheet_name, header=None, nrows=4)
    param_names = df_header.iloc[3].tolist()  # Parameter names (row index 3)
    unit_row = df_header.iloc[2].tolist()     # Units (row index 2)
    
    # Build column->unit mapping from Excel
    col_units = {}
    for i, (param, unit) in enumerate(zip(param_names, unit_row)):
        if pd.notna(param) and pd.notna(unit):
            # Clean unit: remove " TS" suffix (tørrstoff indicator)
            unit_clean = str(unit).strip().replace(' TS', '')
            col_units[param] = unit_clean
    
    # Find Prøvemerking column index for color extraction
    provemerking_col_idx = None
    for i, name in enumerate(param_names):
        if name == 'Prøvemerking':
            provemerking_col_idx = i + 1  # openpyxl is 1-indexed
            break
    
    # Read with header at row 3 (0-indexed), data starts at row 4
    df = pd.read_excel(filepath, sheet_name=sheet_name, header=3)
    
    samples = []
    results = []
    classifications = []
    
    # Find the relevant columns
    # Row 2 has actual column names like 'Prøvemerking', 'Fra pel', etc.
    
    for idx, row in df.iterrows():
        # Skip rows without sample data
        provemerking = row.get('Prøvemerking', None)
        if pd.isna(provemerking) or provemerking == 'Prøvemerking' or str(provemerking).strip() == '':
            continue
        
        # Extract sample metadata
        fra_pel = row.get('Fra pel', None)
        til_pel = row.get('Til Pel', None)
        lab_ref = row.get('Eurofins prøvenummer', '')
        test_utgatt = row.get('Test/utgått', None)
        
        # Skip test/utgått samples
        if test_utgatt == 'x':
            continue
        
        # Generate sample ID - clean any whitespace/newlines
        provemerking_clean = str(provemerking).replace('\n', '').replace('\r', '').replace('_', '-').strip()
        
        # Strip location prefix if provemerking already starts with it (avoid duplication like TVN-TVN-)
        if provemerking_clean.startswith(f"{location_code}-"):
            provemerking_clean = provemerking_clean[len(location_code)+1:]
        
        sample_id = f"p18-{location_code}-{provemerking_clean}"
        
        # Extract tilstandsklasse from cell background color
        tilstandsklasse = None
        if ws_openpyxl and provemerking_col_idx:
            # Excel row = pandas idx + 5 (header rows 1-4, data starts row 5)
            excel_row = idx + 5
            cell = ws_openpyxl.cell(row=excel_row, column=provemerking_col_idx)
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                color = cell.fill.fgColor.rgb
                tilstandsklasse = COLOR_TO_TILSTANDSKLASSE.get(color)
        
        # Clean lab reference (may have newlines from merged cells)
        lab_ref_clean = str(lab_ref).replace('\n', '').replace('\r', '').strip() if not pd.isna(lab_ref) else ''
        
        # Create sample record
        sample = {
            'sample_id': sample_id,
            'project_code': PROJECT_CODE,
            'sample_date': '',  # Not in Excel, would need to get from other docs
            'location_type': sheet_name,
            'profile_start': fra_pel if not pd.isna(fra_pel) else '',
            'profile_end': til_pel if not pd.isna(til_pel) else '',
            'tunnel_name': TUNNEL_NAME,
            'sample_type': 'bunnrensk',
            'lab_reference': lab_ref_clean,
            'sampler': 'Veidekke',
        }
        samples.append(sample)
        
        # Store classification
        classifications.append({
            'sample_id': sample_id,
            'tilstandsklasse': tilstandsklasse,
            'limiting_parameters': '',
            'classification_basis': 'TA-2553/2009' if tilstandsklasse else '',
        })
        
        # Extract all parameter values from columns in EUROFINS_COLUMN_MAP
        for excel_col in df.columns:
            if excel_col in SKIP_COLUMNS:
                continue
            
            # Get parameter code from shared mapping
            param_code = get_parameter_code(excel_col)
            if not param_code or param_code == excel_col:
                # Skip unmapped columns (likely not chemical parameters)
                continue
            
            val = row[excel_col]
            numeric_val, below_limit, loq = parse_value(val)
            
            if numeric_val is not None:
                # Use unit from Excel row 2 if available, otherwise default mg/kg
                unit = col_units.get(excel_col, 'mg/kg')
                results.append({
                    'sample_id': sample_id,
                    'parameter': param_code,
                    'parameter_raw': excel_col,
                    'value': numeric_val,
                    'unit': unit,
                    'uncertainty': None,
                    'below_limit': below_limit,
                    'loq': loq,
                    'analysis_type': 'totalanalyse',
                })
    
    return samples, results, classifications


def create_wide_table(results_df: pd.DataFrame) -> pd.DataFrame:
    """
    Create a wide format table with samples as columns.
    Includes analysis_type to distinguish totalanalyse, ristetest, kolonnetest.
    """
    if len(results_df) == 0:
        return pd.DataFrame()
    
    # Create formatted value strings
    def format_value(row):
        if pd.isna(row['value']):
            return ''
        val = row['value']
        if val == int(val):
            val_str = str(int(val))
        elif val < 0.01:
            val_str = f"{val:.4f}"
        elif val < 1:
            val_str = f"{val:.3f}"
        else:
            val_str = f"{val:.2f}".rstrip('0').rstrip('.')
        
        if row['below_limit']:
            val_str = f"<{val_str}"
        return val_str
    
    results_df = results_df.copy()
    results_df['formatted'] = results_df.apply(format_value, axis=1)
    
    # Get unit per sample
    sample_units = results_df.groupby('sample_id')['unit'].agg(
        lambda x: x.value_counts().index[0]
    ).to_dict()
    
    # Pivot with analysis_type in the index
    wide_df = results_df.pivot_table(
        index=['analysis_type', 'parameter', 'parameter_raw'],
        columns='sample_id',
        values='formatted',
        aggfunc='first'
    ).reset_index()
    
    wide_df.columns.name = None
    wide_df = wide_df.rename(columns={'parameter': 'param_code', 'parameter_raw': 'parameter'})
    
    # Add units to column headers
    sample_cols = [c for c in wide_df.columns if c.startswith('p18-')]
    sample_cols.sort()
    col_rename = {col: f"{col} ({sample_units.get(col, 'mg/kg')})" for col in sample_cols}
    wide_df = wide_df.rename(columns=col_rename)
    
    new_sample_cols = [col_rename[c] for c in sample_cols]
    wide_df = wide_df[['analysis_type', 'param_code', 'parameter'] + new_sample_cols]
    
    return wide_df


def save_wide_table_xlsx(wide_df: pd.DataFrame, filepath: Path) -> None:
    """Save wide table as formatted Excel."""
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    
    for r_idx, row in enumerate(dataframe_to_rows(wide_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    max_row = len(wide_df) + 1
    max_col = len(wide_df.columns)
    table_range = f"A1:{get_column_letter(max_col)}{max_row}"
    
    table = Table(displayName="ResultsTable", ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)
    
    for col_idx, col in enumerate(wide_df.columns, 1):
        max_width = len(str(col))
        for row_idx in range(2, max_row + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_width = max(max_width, len(str(cell_value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 2, 30)
    
    ws.freeze_panes = "A2"
    wb.save(filepath)


def create_qa_workbook(filepath: Path, summary: dict, samples_df: pd.DataFrame,
                       results_df: pd.DataFrame, class_df: pd.DataFrame, 
                       dec_df: pd.DataFrame) -> None:
    """
    Create a QA workbook for manual verification of extracted data.
    
    Sheets:
    - Summary: Extraction metadata and file overview
    - QA Checklist: Items to verify with status columns
    - Samples QA: Sample data with QA columns
    - Classifications QA: Classification data with QA columns
    - Decisions QA: Decision data with QA columns
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # =========================================================================
    # SHEET 1: Summary
    # =========================================================================
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Title
    ws_summary['A1'] = f"QA Workbook: {summary['project']}"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:D1')
    
    # Extraction info
    summary_items = [
        ('Project Code', summary['project_code']),
        ('Extracted At', summary['extracted_at']),
        ('Source File', summary['source_file']),
        ('Lab', summary['lab']),
        ('Sheets Processed', summary['sheets_processed']),
    ]
    
    for i, (label, value) in enumerate(summary_items, start=3):
        ws_summary[f'A{i}'] = label
        ws_summary[f'A{i}'].font = Font(bold=True)
        ws_summary[f'B{i}'] = value
    
    # Data counts
    ws_summary['A10'] = "Data Exports"
    ws_summary['A10'].font = Font(bold=True, size=12)
    
    export_items = [
        ('File', 'Rows', 'Description', 'Remarks', 'Date', 'Name'),
        ('p18_samples.csv', summary['samples_count'], 'Sample metadata', '', '', ''),
        ('p18_results.csv', summary['results_count'], 'Lab analysis results (long format)', '', '', ''),
        ('p18_results_wide.csv', len(results_df['parameter'].unique()) if len(results_df) > 0 else 0, 'Lab results (wide format for review)', '', '', ''),
        ('p18_results_wide.xlsx', len(results_df['parameter'].unique()) if len(results_df) > 0 else 0, 'Lab results (Excel table)', '', '', ''),
        ('p18_classifications.csv', summary['classifications_count'], 'Tilstandsklasse per sample', '', '', ''),
        ('p18_decisions.csv', summary['decisions_count'], 'Handling decisions per sample', '', '', ''),
        ('p18_extraction_summary.csv', 1, 'Extraction metadata', '', '', ''),
    ]
    
    for i, row_data in enumerate(export_items, start=11):
        for j, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=i, column=j, value=value)
            if i == 11:  # Header row
                cell.font = header_font
                cell.fill = header_fill
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 10
    ws_summary.column_dimensions['C'].width = 40
    ws_summary.column_dimensions['D'].width = 30
    ws_summary.column_dimensions['E'].width = 12
    ws_summary.column_dimensions['F'].width = 15
    
    # =========================================================================
    # SHEET 2: QA Checklist
    # =========================================================================
    ws_qa = wb.create_sheet("QA Checklist")
    
    checklist_items = [
        ('Category', 'Check Item', 'Status', 'Verified By', 'Date', 'Notes'),
        ('Samples', 'All samples from source document captured', '', '', '', ''),
        ('Samples', 'Sample IDs match lab report numbers', '', '', '', ''),
        ('Samples', 'Profile locations match source', '', '', '', ''),
        ('Results', 'All parameters extracted for each sample', '', '', '', ''),
        ('Results', 'Values match source Excel', '', '', '', ''),
        ('Results', 'Below-limit values (<) correctly flagged', '', '', '', ''),
        ('Results', 'Units correct (mg/kg, µg/kg, %)', '', '', '', ''),
        ('Classifications', 'Tilstandsklasse matches source', '', '', '', ''),
        ('Classifications', 'Limiting parameters identified correctly', '', '', '', ''),
        ('Decisions', 'Handling decisions match documentation', '', '', '', ''),
        ('Decisions', 'Destinations correct (deponi, gjenbruk, etc)', '', '', '', ''),
        ('General', 'No duplicate samples', '', '', '', ''),
        ('General', 'No missing critical data', '', '', '', ''),
        ('General', 'Data ready for aggregation', '', '', '', ''),
    ]
    
    for i, row in enumerate(checklist_items, start=1):
        for j, value in enumerate(row, start=1):
            cell = ws_qa.cell(row=i, column=j, value=value)
            cell.border = thin_border
            if i == 1:  # Header
                cell.font = header_font
                cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Set column widths
    ws_qa.column_dimensions['A'].width = 15
    ws_qa.column_dimensions['B'].width = 45
    ws_qa.column_dimensions['C'].width = 12
    ws_qa.column_dimensions['D'].width = 15
    ws_qa.column_dimensions['E'].width = 12
    ws_qa.column_dimensions['F'].width = 40
    
    # Freeze header
    ws_qa.freeze_panes = 'A2'
    
    # =========================================================================
    # SHEET 3: Samples QA
    # =========================================================================
    ws_samples = wb.create_sheet("Samples QA")
    
    # Add QA columns to samples
    samples_qa = samples_df.copy()
    samples_qa['QA_Status'] = ''
    samples_qa['QA_Notes'] = ''
    
    for i, row in enumerate(dataframe_to_rows(samples_qa, index=False, header=True), start=1):
        for j, value in enumerate(row, start=1):
            cell = ws_samples.cell(row=i, column=j, value=value)
            if i == 1:
                cell.font = header_font
                cell.fill = header_fill
    
    ws_samples.freeze_panes = 'A2'
    
    # Auto-width columns
    for col_idx, col in enumerate(samples_qa.columns, start=1):
        ws_samples.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col)) + 2, 12)
    
    # =========================================================================
    # SHEET 4: Classifications QA
    # =========================================================================
    ws_class = wb.create_sheet("Classifications QA")
    
    class_qa = class_df.copy()
    class_qa['QA_Status'] = ''
    class_qa['QA_Notes'] = ''
    
    for i, row in enumerate(dataframe_to_rows(class_qa, index=False, header=True), start=1):
        for j, value in enumerate(row, start=1):
            cell = ws_class.cell(row=i, column=j, value=value)
            if i == 1:
                cell.font = header_font
                cell.fill = header_fill
    
    ws_class.freeze_panes = 'A2'
    
    for col_idx, col in enumerate(class_qa.columns, start=1):
        ws_class.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col)) + 2, 15)
    
    # =========================================================================
    # SHEET 5: Decisions QA
    # =========================================================================
    ws_dec = wb.create_sheet("Decisions QA")
    
    dec_qa = dec_df.copy()
    dec_qa['QA_Status'] = ''
    dec_qa['QA_Notes'] = ''
    
    for i, row in enumerate(dataframe_to_rows(dec_qa, index=False, header=True), start=1):
        for j, value in enumerate(row, start=1):
            cell = ws_dec.cell(row=i, column=j, value=value)
            if i == 1:
                cell.font = header_font
                cell.fill = header_fill
    
    ws_dec.freeze_panes = 'A2'
    
    for col_idx, col in enumerate(dec_qa.columns, start=1):
        ws_dec.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col)) + 2, 15)
    
    # Save workbook
    wb.save(filepath)


def extract():
    """Main extraction function for Project 18."""
    from openpyxl import load_workbook
    
    print(f"{'='*60}")
    print(f"Extracting: {PROJECT_NAME}")
    print(f"{'='*60}")
    
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    excel_path = INBOX_DIR / EXCEL_FILE
    
    if not excel_path.exists():
        print(f"ERROR: Excel file not found: {excel_path}")
        return False
    
    print(f"\nSource: {EXCEL_FILE}")
    
    # Load workbook with openpyxl for color extraction
    wb_openpyxl = load_workbook(excel_path, data_only=True)
    
    # ============================================================
    # READ ALL SHEETS
    # ============================================================
    all_samples = []
    all_results = []
    all_classifications = []
    
    for sheet_name, location_code in SHEETS.items():
        print(f"\n[Reading] {sheet_name} ({location_code})...")
        try:
            ws = wb_openpyxl[sheet_name]
            samples, results, classifications = read_sheet(excel_path, sheet_name, location_code, ws)
            all_samples.extend(samples)
            all_results.extend(results)
            all_classifications.extend(classifications)
            print(f"  Found {len(samples)} samples, {len(results)} results")
        except Exception as e:
            print(f"  ERROR reading sheet: {e}")
    
    # ============================================================
    # READ LEACHING TEST RESULTS FROM PDF (pages 34-37)
    # ============================================================
    pdf_path = INBOX_DIR / PDF_FILE
    if pdf_path.exists():
        print(f"\n[Reading] Leaching tests from PDF (pages {PDF_LEACHING_PAGES[0]}-{PDF_LEACHING_PAGES[-1]})...")
        try:
            leaching_results = parse_leaching_test_pdf(pdf_path, PDF_LEACHING_PAGES)
            all_results.extend(leaching_results)
            print(f"  Found {len(leaching_results)} leaching test results")
        except Exception as e:
            print(f"  ERROR reading PDF: {e}")
    else:
        print(f"\n[Skipping] PDF not found: {PDF_FILE}")
    
    # ============================================================
    # BUILD SAMPLES TABLE
    # ============================================================
    print(f"\n[1/4] Building p18_samples.csv...")
    
    samples_df = pd.DataFrame(all_samples)
    save_to_csv(samples_df, OUTPUT_DIR / 'p18_samples.csv')
    print(f"  Saved: p18_samples.csv ({len(samples_df)} samples)")
    
    # ============================================================
    # BUILD RESULTS TABLE
    # ============================================================
    print(f"\n[2/4] Building p18_results.csv...")
    
    results_df = pd.DataFrame(all_results)
    if len(results_df) > 0:
        results_df = results_df[['sample_id', 'parameter', 'parameter_raw', 'value', 
                                  'unit', 'uncertainty', 'below_limit', 'loq', 'analysis_type']]
        save_to_csv(results_df, OUTPUT_DIR / 'p18_results.csv')
        print(f"  Saved: p18_results.csv ({len(results_df)} results)")
        
        # Summary by sample
        sample_counts = results_df.groupby('sample_id').size()
        for sid, count in list(sample_counts.items())[:5]:
            print(f"    {sid}: {count} parameters")
        if len(sample_counts) > 5:
            print(f"    ... and {len(sample_counts) - 5} more samples")
        
        # Wide format
        wide_df = create_wide_table(results_df)
        save_to_csv(wide_df, OUTPUT_DIR / 'p18_results_wide.csv')
        print(f"  Saved: p18_results_wide.csv ({len(wide_df)} parameters x {len(wide_df.columns)-2} samples)")
        
        save_wide_table_xlsx(wide_df, OUTPUT_DIR / 'p18_results_wide.xlsx')
        print(f"  Saved: p18_results_wide.xlsx")
    else:
        print("  WARNING: No results found!")
        results_df = pd.DataFrame()
    
    # ============================================================
    # BUILD CLASSIFICATIONS (from cell colors)
    # ============================================================
    print(f"\n[3/4] Building p18_classifications.csv...")
    
    class_df = pd.DataFrame(all_classifications)
    save_to_csv(class_df, OUTPUT_DIR / 'p18_classifications.csv')
    
    # Count how many have tilstandsklasse extracted
    has_class = class_df['tilstandsklasse'].notna().sum()
    print(f"  Saved: p18_classifications.csv ({len(class_df)} classifications)")
    print(f"  Extracted tilstandsklasse from cell colors: {has_class}/{len(class_df)}")
    
    # ============================================================
    # BUILD DECISIONS (based on tilstandsklasse)
    # ============================================================
    print(f"\n[4/4] Building p18_decisions.csv...")
    
    # Create lookup from sample_id to tilstandsklasse
    class_lookup = {c['sample_id']: c['tilstandsklasse'] for c in all_classifications}
    
    dec_data = []
    for s in all_samples:
        sample_id = s['sample_id']
        tilstandsklasse = class_lookup.get(sample_id)
        
        # Tilstandsklasse 1 = gjenbruk, otherwise deponi
        if tilstandsklasse == 1:
            decision = 'under normverdier'
            destination = 'gjenbruk'
        elif tilstandsklasse is not None:
            decision = 'over normverdier'
            destination = 'deponi'
        else:
            decision = 'ukjent'
            destination = ''
        
        dec_data.append({
            'sample_id': sample_id,
            'decision': decision,
            'destination': destination,
            'notes': 'Autogenerert basert på tilstandsklasse' if tilstandsklasse else 'Tilstandsklasse ikke satt',
        })
    
    dec_df = pd.DataFrame(dec_data)
    save_to_csv(dec_df, OUTPUT_DIR / 'p18_decisions.csv')
    print(f"  Saved: p18_decisions.csv ({len(dec_df)} decisions)")
    
    # ============================================================
    # EXTRACTION SUMMARY
    # ============================================================
    summary = {
        'project': PROJECT_NAME,
        'project_code': PROJECT_CODE,
        'extracted_at': datetime.now().isoformat(),
        'source_file': EXCEL_FILE,
        'samples_count': len(samples_df),
        'results_count': len(results_df),
        'classifications_count': len(class_df),
        'decisions_count': len(dec_df),
        'lab': 'Eurofins',
        'sheets_processed': ', '.join(SHEETS.keys()),
    }
    
    summary_df = pd.DataFrame([summary])
    save_to_csv(summary_df, OUTPUT_DIR / 'p18_extraction_summary.csv')
    
    # ============================================================
    # QA WORKBOOK
    # ============================================================
    qa_workbook_path = OUTPUT_DIR / 'p18_QA_workbook.xlsx'
    if qa_workbook_path.exists():
        # Rename existing file with timestamp to preserve manual notes
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = OUTPUT_DIR / f'p18_QA_workbook_{timestamp}.xlsx'
        qa_workbook_path.rename(backup_path)
        print(f"  Renamed existing QA workbook to: p18_QA_workbook_{timestamp}.xlsx")
    
    create_qa_workbook(
        qa_workbook_path,
        summary,
        samples_df,
        results_df,
        class_df,
        dec_df
    )
    print(f"  Saved: p18_QA_workbook.xlsx (for manual QA documentation)")
    
    # ============================================================
    # FINAL REPORT
    # ============================================================
    print(f"\n{'='*60}")
    print("EXTRACTION COMPLETE")
    print(f"{'='*60}")
    print(f"\nOutput files in: {OUTPUT_DIR}")
    print(f"  - p18_samples.csv          ({len(samples_df)} rows)")
    print(f"  - p18_results.csv          ({len(results_df)} rows)")
    print(f"  - p18_classifications.csv  ({len(class_df)} rows)")
    print(f"  - p18_decisions.csv        ({len(dec_df)} rows)")
    print(f"  - p18_extraction_summary.csv")
    print(f"  - p18_QA_workbook.xlsx")
    print(f"{'='*60}")
    
    return True


if __name__ == '__main__':
    success = extract()
    sys.exit(0 if success else 1)
