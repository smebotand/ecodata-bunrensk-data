"""
Project 02 - E18 Vestkorridoren E103 (Strand - Ramstadsletta) Extraction Script

Extracts data from ALS Excel export with multiple samples in long format.
Source: SVV E18 Vestkorridoren E103 - Alle bunnrenskprøver.xlsx

Tunnels:
- Høviktoppen (various profiles)
- Various pel. locations

Output:
- p02_samples.csv
- p02_results.csv
- p02_classifications.csv
- p02_decisions.csv
- p02_results_wide.csv
- p02_results_wide.xlsx
"""

import sys
import re
from pathlib import Path
from datetime import datetime
import pandas as pd
import openpyxl

# Add parent to path for lib imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from lib.export import save_to_csv
from lib.chemistry import normalize_parameter_name
from lib.excel_utils import create_wide_table, save_wide_table_xlsx
from lib.qa import create_qa_workbook
from lib.dataframes import (
    create_samples_df, create_results_df,
    create_classifications_df, create_decisions_df
)

# ============================================================
# PROJECT CONFIGURATION
# ============================================================

PROJECT_CODE = '02_e18-e103'
PROJECT_NAME = 'E18 Vestkorridoren E103 - Strand Ramstadsletta'
TUNNEL_NAME = 'E103 Høviktoppen'
PROJECT_PREFIX = 'p02-'

BASE_DIR = Path(__file__).parent.parent.parent
INBOX_DIR = BASE_DIR / '00_inbox' / 'SVV' / 'E18 Vestkorridoren E103'
RESULTATER_DIR = INBOX_DIR / 'resultater'
OUTPUT_DIR = BASE_DIR / '01_projects' / '02_e18-e103' / 'extracted'

EXCEL_FILE = 'Alle bunnrenskprøver, mangler siste blandeprøve ØL (1).xlsx'
SHEET_NAME = 'Resultater med grenseverdier -1'

# Column indices (1-based for openpyxl)
COL_ORDER_NR = 6       # Order number (e.g., NO2522963)
COL_SAMPLE_NAME = 7    # Sample name
COL_SAMPLE_DESC = 8    # Sample description
COL_SAMPLE_DATE = 9    # Prøvetakingsdato
COL_PRØVE_NR = 16      # Prøve nr. (e.g., NO2500215-005)
COL_ANALYSE = 18       # Analyse (parameter name)
COL_UNIT = 19          # Enhet
COL_RESULT = 20        # Resultat
COL_GRENSEVERDI = 23   # Grenseverdi (Tilstandsklasse name)
COL_EVAL = 24          # Evaluering (Godkjent/Ikke-godkjent)

# Filter to include only bunnrensk samples
BUNNRENSK_KEYWORDS = ['bunnrensk', 'blandeprøve pel', 'pel.', 'pel ', 'grøftekant', 'grøfteslam', 'brmø']

# Parameters to skip
SKIP_PARAMS = {'Tørrstoff ved 105 grader', 'Vanninnhold', 'TOC'}

# Tilstandsklasse from grenseverdi text
GRENSE_TO_TK = {
    'Tilstandsklasse 1- Meget god': 1,
    'Tilstandsklasse 2 - God': 2,
    'Tilstandsklasse 3 - Moderat': 3,
    'Tilstandsklasse 4 - Dårlig': 4,
    'Tillstandsklasse 5 - Svært dårlig': 5,  # Note: typo in Excel "Tillstandsklasse"
    'Tilstandsklasse 5 - Svært dårlig': 5,
}


def save_source_samples_xlsx(df: pd.DataFrame, filepath: Path) -> None:
    """
    Save source samples as formatted Excel table.
    """
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Source Samples"
    
    # Write dataframe to worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Define table range
    max_row = len(df) + 1  # +1 for header
    max_col = len(df.columns)
    table_range = f"A1:{get_column_letter(max_col)}{max_row}"
    
    # Create table with style
    table = Table(displayName="SourceSamplesTable", ref=table_range)
    style = TableStyleInfo(
        name="TableStyleMedium2",  # Blue alternating rows
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    
    # Auto-adjust column widths
    for col_idx, col in enumerate(df.columns, 1):
        max_width = len(str(col))  # Header width
        for row_idx in range(2, min(max_row + 1, 102)):  # Sample first 100 rows
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_width = max(max_width, len(str(cell_value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 2, 50)
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    wb.save(filepath)


def is_bunnrensk_sample(sample_name: str) -> bool:
    """Check if sample name indicates a bunnrensk sample."""
    if not sample_name:
        return False
    name_lower = sample_name.lower()
    return any(kw in name_lower for kw in BUNNRENSK_KEYWORDS)


def scan_resultater_folder() -> list:
    """
    Scan all Excel files in the resultater folder for sample names.
    Sample names are in row 7, starting from column C (3) horizontally.
    
    Returns list of dicts with sample info.
    """
    samples = []
    
    if not RESULTATER_DIR.exists():
        print(f"  Resultater folder not found: {RESULTATER_DIR}")
        return samples
    
    xlsx_files = list(RESULTATER_DIR.glob('*.xlsx'))
    print(f"  Scanning {len(xlsx_files)} Excel files in resultater folder...")
    
    seen_samples = set()
    
    for f in xlsx_files:
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
            
            # Extract order number from filename (e.g., EDD_XLS_STD_NO2326608_0_nb-NO.xlsx)
            order_match = re.search(r'NO\d+', f.name)
            order_nr = order_match.group(0) if order_match else ''
            
            # Get samples from row 7, starting col 3
            for col in range(3, 30):
                val = ws.cell(row=7, column=col).value
                if val and val not in ['ELEMENT', 'SAMPLE', '']:
                    sample_name = str(val).strip()
                    if sample_name and sample_name not in seen_samples:
                        seen_samples.add(sample_name)
                        is_included = is_bunnrensk_sample(sample_name)
                        samples.append({
                            'source_sample_name': sample_name,
                            'sample_description': '',  # Not available in this format
                            'sample_date': '',  # Not available in row 7
                            'order_nr': order_nr,
                            'source_file': f.name,
                            'included': 'Yes' if is_included else 'No',
                            'matched_keyword': next((kw for kw in BUNNRENSK_KEYWORDS if kw in sample_name.lower()), '') if is_included else '',
                        })
            wb.close()
        except Exception as e:
            print(f"    Warning: Could not read {f.name}: {e}")
    
    print(f"  Found {len(samples)} unique samples in resultater folder")
    return samples


def parse_value(val):
    """
    Parse a value from the Excel file.
    Returns (numeric_value, below_limit, loq)
    """
    if pd.isna(val) or val == '' or val == 'nd':
        return None, False, None
    
    val_str = str(val).strip()
    
    # Handle below-limit values like "< 0,20" or "<0.20" or "<2.0"
    if val_str.startswith('<'):
        val_str = val_str[1:].strip()
        val_str = val_str.replace(',', '.').replace(' ', '')
        try:
            loq = float(val_str)
            return loq, True, loq
        except:
            return None, True, None
    
    # Regular numeric value
    val_str = val_str.replace(',', '.')
    try:
        return float(val_str), False, None
    except:
        return None, False, None


def parse_profile_from_sample_name(sample_name: str) -> tuple:
    """Extract profile start/end from sample name.
    
    Examples:
    - 'Blandeprøve pel 9260-9110' -> (9110, 9260)
    - 'Bunnrenskprøve pel. 9810' -> (9810, None)
    - 'Bunnrensk 8760' -> (8760, None)
    - 'Bunnrensk Høviktoppen a 0,5m' -> (None, None)
    """
    if not sample_name:
        return None, None
    
    # Look for patterns like 9260-9110 or 9260 - 9110 or pel. 9810
    numbers = re.findall(r'(\d{4,5})', str(sample_name))
    if len(numbers) >= 2:
        # Sort to get correct order (start < end)
        nums = sorted([float(n) for n in numbers])
        return nums[0], nums[-1]
    elif len(numbers) == 1:
        return float(numbers[0]), None
    return None, None


def get_location_type(sample_name: str) -> str:
    """Determine location_type from sample name."""
    name_lower = sample_name.lower()
    if 'grøft' in name_lower:
        return 'grøft'
    elif 'pumpesump' in name_lower:
        return 'pumpesump'
    else:
        return 'vegbane'


def get_sample_type(sample_name: str) -> str:
    """Determine sample_type from sample name."""
    name_lower = sample_name.lower()
    if 'blandeprøve' in name_lower:
        return 'blandprøve'
    else:
        return 'bunnrensk'


def get_analysis_type_from_sample_name(sample_name: str) -> str:
    """Determine analysis_type from sample name suffix.
    
    Some samples have analysis type encoded in name:
    - 'BRMØ1' -> totalanalyse
    - 'BRMØ1 ristetest' -> ristetest
    - 'BRMØ1 kolonnetest' -> kolonnetest
    """
    name_lower = sample_name.lower()
    if 'kolonnetest' in name_lower:
        return 'kolonnetest'
    elif 'ristetest' in name_lower:
        return 'ristetest'
    else:
        return 'totalanalyse'


def get_base_sample_name(sample_name: str) -> str:
    """Get base sample name without analysis type suffix.
    
    - 'BRMØ1 kolonnetest' -> 'BRMØ1'
    - 'BRMØ1 ristetest' -> 'BRMØ1'
    - 'Bunnrensk 8760' -> 'Bunnrensk 8760' (unchanged)
    """
    # Remove analysis type suffixes
    name = sample_name
    for suffix in [' kolonnetest', ' ristetest', ' Kolonnetest', ' Ristetest']:
        if name.endswith(suffix):
            name = name[:-len(suffix)]
            break
    return name


def extract_data(excel_path: Path) -> tuple[list, list, list]:
    """
    Extract samples and results from the E103 Excel file.
    
    The Excel file has one row per (sample, parameter, grenseverdi) combination.
    We filter to only TK1 grenseverdi rows to get one result per (sample, parameter).
    
    Returns:
        Tuple of (samples_list, results_list, all_source_samples_list)
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[SHEET_NAME]
    
    print(f"Total rows in sheet: {ws.max_row}")
    
    # First pass: collect ALL unique samples (for source overview)
    all_samples_dict = {}  # sample_name -> sample info with included flag
    samples_dict = {}  # sample_name -> sample info (only bunnrensk)
    
    for row in range(3, ws.max_row + 1):
        sample_name = ws.cell(row=row, column=COL_SAMPLE_NAME).value
        if not sample_name:
            continue
        # Exclude specific sample by name (case-insensitive, ignore spaces and punctuation)
        name_clean = str(sample_name).lower().replace(' ', '').replace('-', '').replace('.', '')
        if name_clean in [
            'sprengsteinpel9160',
            'p02sprengsteinpel9160',
            'sprengstein-pel-9160',
            'p02-sprengstein-pel-9160',
        ]:
            continue
        
        # Track ALL samples for source overview
        if sample_name not in all_samples_dict:
            order_nr = ws.cell(row=row, column=COL_ORDER_NR).value
            sample_date = ws.cell(row=row, column=COL_SAMPLE_DATE).value
            sample_desc = ws.cell(row=row, column=COL_SAMPLE_DESC).value
            
            if isinstance(sample_date, datetime):
                sample_date_str = sample_date.strftime('%Y-%m-%d')
            elif hasattr(sample_date, 'strftime'):
                sample_date_str = sample_date.strftime('%Y-%m-%d')
            else:
                sample_date_str = ''
            
            is_included = is_bunnrensk_sample(sample_name)
            all_samples_dict[sample_name] = {
                'source_sample_name': sample_name,
                'sample_description': str(sample_desc).strip() if sample_desc else '',
                'sample_date': sample_date_str,
                'order_nr': str(order_nr).strip() if order_nr else '',
                'source_file': EXCEL_FILE,
                'included': 'Yes' if is_included else 'No',
                'matched_keyword': next((kw for kw in BUNNRENSK_KEYWORDS if kw in sample_name.lower()), '') if is_included else '',
            }
        
        # Skip non-bunnrensk for main extraction
        if not is_bunnrensk_sample(sample_name):
            continue
        
        if sample_name not in samples_dict:
            order_nr = ws.cell(row=row, column=COL_ORDER_NR).value
            sample_date = ws.cell(row=row, column=COL_SAMPLE_DATE).value
            
            # Convert date
            if isinstance(sample_date, datetime):
                sample_date_str = sample_date.strftime('%Y-%m-%d')
            elif hasattr(sample_date, 'strftime'):
                sample_date_str = sample_date.strftime('%Y-%m-%d')
            else:
                sample_date_str = ''
            
            # Parse profile from sample name
            profile_start, profile_end = parse_profile_from_sample_name(sample_name)
            
            # Create sample ID
            clean_name = sample_name.replace(' ', '-').replace(',', '').replace('.', '')
            sample_id = f"{PROJECT_PREFIX}{clean_name}"
            
            samples_dict[sample_name] = {
                'sample_id': sample_id,
                'project_code': PROJECT_CODE,
                'sample_date': sample_date_str,
                'location_type': get_location_type(sample_name),
                'profile_start': profile_start,
                'profile_end': profile_end,
                'tunnel_name': TUNNEL_NAME,
                'sample_type': get_sample_type(sample_name),
                'lab_reference': str(order_nr).strip() if order_nr else '',
                'sampler': 'SVV',
                'remark': '',
            }
    
    print(f"Found {len(samples_dict)} bunnrensk samples")
    
    # Second pass: collect results (only TK1 rows to avoid duplicates)
    results_dict = {}  # (sample_name, analyse) -> result info
    
    for row in range(3, ws.max_row + 1):
        sample_name = ws.cell(row=row, column=COL_SAMPLE_NAME).value
        if not sample_name or not is_bunnrensk_sample(sample_name):
            continue
        # Exclude specific sample by name (case-insensitive, ignore spaces and punctuation)
        name_clean = str(sample_name).lower().replace(' ', '').replace('-', '').replace('.', '')
        if name_clean in [
            'sprengsteinpel9160',
            'p02sprengsteinpel9160',
            'sprengstein-pel-9160',
            'p02-sprengstein-pel-9160',
        ]:
            continue
        
        grenseverdi = ws.cell(row=row, column=COL_GRENSEVERDI).value
        if grenseverdi != 'Tilstandsklasse 1- Meget god':
            continue  # Only process TK1 rows to get one row per (sample, analyse)
        
        analyse = ws.cell(row=row, column=COL_ANALYSE).value
        if not analyse or analyse in SKIP_PARAMS:
            continue
        
        result_val = ws.cell(row=row, column=COL_RESULT).value
        value, below_limit, loq = parse_value(result_val)
        
        if value is None:
            continue
        
        unit = ws.cell(row=row, column=COL_UNIT).value or 'mg/kg'
        unit = str(unit).replace(' TS', '').strip()  # Clean up "mg/kg TS" -> "mg/kg"
        
        # Determine tilstandsklasse based on evaluation
        eval_result = ws.cell(row=row, column=COL_EVAL).value
        tilstandsklasse = 1 if eval_result == 'Godkjent' else None
        
        # For non-godkjent, we need to check higher TK levels
        # But since we filter TK1, "Ikke-godkjent" means TK > 1
        
        key = (sample_name, analyse)
        if key not in results_dict:
            sample_id = samples_dict[sample_name]['sample_id']
            param_code = normalize_parameter_name(analyse)
            
            results_dict[key] = {
                'sample_id': sample_id,
                'parameter': param_code,
                'parameter_raw': analyse,
                'value': value,
                'unit': unit,
                'uncertainty': None,
                'below_limit': below_limit,
                'loq': loq,
                'analysis_type': get_analysis_type_from_sample_name(sample_name),
                'row': row,  # Keep track for TK lookup
            }
    
    print(f"Found {len(results_dict)} unique results")
    
    # Third pass: determine tilstandsklasse for each result
    # TK is the lowest TK where "Godkjent"
    # If TK1=Godkjent -> TK=1
    # If TK1=Ikke-godkjent, TK2=Godkjent -> TK=2
    # etc.
    # If "Ingen spesifikasjon" on all -> TK=None (no limit for this parameter)
    results_with_tk_final = {}
    
    for row in range(3, ws.max_row + 1):
        sample_name = ws.cell(row=row, column=COL_SAMPLE_NAME).value
        if not sample_name or not is_bunnrensk_sample(sample_name):
            continue
        
        analyse = ws.cell(row=row, column=COL_ANALYSE).value
        if not analyse or analyse in SKIP_PARAMS:
            continue
        
        key = (sample_name, analyse)
        if key not in results_dict:
            continue
        
        grenseverdi = ws.cell(row=row, column=COL_GRENSEVERDI).value
        eval_result = ws.cell(row=row, column=COL_EVAL).value
        tk = GRENSE_TO_TK.get(grenseverdi)
        
        if tk and eval_result == 'Godkjent':
            # The lowest TK where approved is the actual TK
            if key not in results_with_tk_final or tk < results_with_tk_final[key]:
                results_with_tk_final[key] = tk
    
    # For results with "Ingen spesifikasjon" (no limit), don't set TK
    # For results never approved at any TK, they exceed TK5 (use None or could use 6)
    
    # Add tilstandsklasse to results
    for key, result in results_dict.items():
        result['tilstandsklasse'] = results_with_tk_final.get(key)
    
    samples_list = list(samples_dict.values())
    results_list = list(results_dict.values())
    all_source_samples_list = list(all_samples_dict.values())
    
    # Remove 'row' key from results (was just for tracking)
    for r in results_list:
        r.pop('row', None)
    
    return samples_list, results_list, all_source_samples_list


def main():
    """Main extraction function"""
    print(f"=" * 60)
    print(f"Project: {PROJECT_NAME}")
    print(f"=" * 60)
    
    excel_path = INBOX_DIR / EXCEL_FILE
    print(f"\nSource file: {excel_path}")
    
    if not excel_path.exists():
        print(f"ERROR: Excel file not found: {excel_path}")
        return
    
    # Create output directory
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    # Extract data
    print("\nExtracting data from Excel...")
    samples_list, results_list, all_source_samples_list = extract_data(excel_path)
    
    # Create DataFrames with standard schema
    samples_df = create_samples_df(samples_list)
    
    # Results need tilstandsklasse temporarily for classification
    results_df_full = pd.DataFrame(results_list)
    results_df = create_results_df(results_list)
    
    print(f"\nExtracted:")
    print(f"  - {len(samples_df)} samples")
    print(f"  - {len(results_df)} results")
    
    # Show sample breakdown by location
    if len(samples_df) > 0:
        print(f"\nSamples by location:")
        for loc, count in samples_df['location_type'].value_counts().items():
            print(f"  - {loc}: {count}")
        
        print(f"\nSamples by type:")
        for st, count in samples_df['sample_type'].value_counts().items():
            print(f"  - {st}: {count}")
    
    # Create classifications and decisions using schema functions
    classifications_df = create_classifications_df(samples_df, results_df_full)
    decisions_df = create_decisions_df(samples_df, results_df_full)
    decisions_df['notes'] = 'Bestemmes på et senere tidspunkt'
    
    # Show classification summary
    if len(classifications_df) > 0:
        print(f"\nSample tilstandsklasse (based on highest per sample):")
        tk_counts = classifications_df['tilstandsklasse'].value_counts().sort_index()
        for tk, count in tk_counts.items():
            if pd.notna(tk):
                print(f"  - TK{int(tk)}: {count} samples")
    
    if len(decisions_df) > 0:
        print(f"\nDecision summary:")
        decision_counts = decisions_df['decision'].value_counts()
        for decision, count in decision_counts.items():
            print(f"  - {decision}: {count} samples")
    
    # Create wide table
    print("\nCreating wide table...")
    wide_df = create_wide_table(results_df, PROJECT_PREFIX)
    print(f"  - {len(wide_df)} parameters in wide table")
    
    # Save outputs
    print("\nSaving outputs...")
    save_to_csv(samples_df, OUTPUT_DIR / 'p02_samples.csv')
    save_to_csv(results_df, OUTPUT_DIR / 'p02_results.csv')
    save_to_csv(classifications_df, OUTPUT_DIR / 'p02_classifications.csv')
    save_to_csv(decisions_df, OUTPUT_DIR / 'p02_decisions.csv')
    save_to_csv(wide_df, OUTPUT_DIR / 'p02_results_wide.csv')
    
    # Also save wide table as Excel
    wide_df.to_excel(OUTPUT_DIR / 'p02_results_wide.xlsx', index=False)
    
    # Create extraction summary for QA workbook
    summary = {
        'project': PROJECT_NAME,
        'project_code': PROJECT_CODE,
        'extracted_at': datetime.now().isoformat(),
        'source_file': EXCEL_FILE,
        'samples_count': len(samples_df),
        'results_count': len(results_df),
        'classifications_count': len(classifications_df),
        'decisions_count': len(decisions_df),
        'lab': 'ALS',
    }
    
    # Create QA workbook (always with timestamp to preserve history)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    qa_workbook_path = OUTPUT_DIR / f'p02_QA_workbook_{timestamp}.xlsx'
    create_qa_workbook(qa_workbook_path, summary, samples_df, results_df, 
                       classifications_df, decisions_df)
    print(f"  Saved: p02_QA_workbook_{timestamp}.xlsx")
    
    # Scan resultater folder for additional samples
    print("\nScanning resultater folder...")
    resultater_samples = scan_resultater_folder()
    
    # Combine all source samples from both sources
    # Use sample name as key to avoid duplicates
    all_samples_combined = {s['source_sample_name']: s for s in all_source_samples_list}
    for s in resultater_samples:
        if s['source_sample_name'] not in all_samples_combined:
            all_samples_combined[s['source_sample_name']] = s
    
    # Save all source samples overview
    all_source_df = pd.DataFrame(list(all_samples_combined.values()))
    all_source_df = all_source_df.sort_values(['included', 'sample_date', 'source_sample_name'], ascending=[False, False, True])
    save_source_samples_xlsx(all_source_df, OUTPUT_DIR / 'p02_all_source_samples.xlsx')
    included_count = len(all_source_df[all_source_df['included'] == 'Yes'])
    excluded_count = len(all_source_df[all_source_df['included'] == 'No'])
    print(f"  - All source samples: {len(all_source_df)} total ({included_count} included, {excluded_count} excluded)")
    print(f"    - From main Excel: {len(all_source_samples_list)}")
    print(f"    - From resultater folder: {len(resultater_samples)} (unique additions: {len(all_samples_combined) - len(all_source_samples_list)})")
    
    print(f"\nDone! Output saved to {OUTPUT_DIR}")


if __name__ == '__main__':
    main()
