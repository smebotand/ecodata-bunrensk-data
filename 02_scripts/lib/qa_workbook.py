"""
QA Workbook utilities for bunnrensk data extraction projects.

Creates Excel workbooks for manual verification of extracted data.
"""

from pathlib import Path
import pandas as pd

# Import schema for validation
from lib.schema import SAMPLE_TYPES, LOCATION_TYPES, ANALYSIS_TYPES, DECISION_TYPES, UNIT_TYPES, LAB_TYPES


def _find_schema_violations(samples_df: pd.DataFrame, results_df: pd.DataFrame,
                            dec_df: pd.DataFrame) -> list[dict]:
    """
    Check data against schema constraints and return list of violations.
    
    Returns:
        List of dicts with keys: table, sample_id, field, value, valid_values
    """
    violations = []
    
    # Check samples for invalid location_type
    if len(samples_df) > 0 and 'location_type' in samples_df.columns:
        for _, row in samples_df.iterrows():
            val = row.get('location_type')
            if val and val not in LOCATION_TYPES:
                violations.append({
                    'table': 'samples',
                    'sample_id': row.get('sample_id', '?'),
                    'field': 'location_type',
                    'value': val,
                    'valid_values': ', '.join(sorted(LOCATION_TYPES)),
                })
    
    # Check samples for invalid sample_type
    if len(samples_df) > 0 and 'sample_type' in samples_df.columns:
        for _, row in samples_df.iterrows():
            val = row.get('sample_type')
            if val and val not in SAMPLE_TYPES:
                violations.append({
                    'table': 'samples',
                    'sample_id': row.get('sample_id', '?'),
                    'field': 'sample_type',
                    'value': val,
                    'valid_values': ', '.join(sorted(SAMPLE_TYPES)),
                })
    
    # Check results for invalid analysis_type
    if len(results_df) > 0 and 'analysis_type' in results_df.columns:
        for _, row in results_df.iterrows():
            val = row.get('analysis_type')
            if val and val not in ANALYSIS_TYPES:
                violations.append({
                    'table': 'results',
                    'sample_id': row.get('sample_id', '?'),
                    'field': 'analysis_type',
                    'value': val,
                    'valid_values': ', '.join(sorted(ANALYSIS_TYPES)),
                })
    
    # Check results for invalid unit
    if len(results_df) > 0 and 'unit' in results_df.columns:
        # Only report unique violations per unit value
        seen_units = set()
        for _, row in results_df.iterrows():
            val = row.get('unit')
            if val and val not in UNIT_TYPES and val not in seen_units:
                seen_units.add(val)
                violations.append({
                    'table': 'results',
                    'sample_id': '(multiple)',
                    'field': 'unit',
                    'value': val,
                    'valid_values': ', '.join(sorted(UNIT_TYPES)),
                })
    
    # Check decisions for invalid decision
    if len(dec_df) > 0 and 'decision' in dec_df.columns:
        for _, row in dec_df.iterrows():
            val = row.get('decision')
            if val and val not in DECISION_TYPES:
                violations.append({
                    'table': 'decisions',
                    'sample_id': row.get('sample_id', '?'),
                    'field': 'decision',
                    'value': val,
                    'valid_values': ', '.join(sorted(DECISION_TYPES)),
                })
    
    return violations


def create_qa_workbook(filepath: Path, summary: dict, samples_df: pd.DataFrame,
                       results_df: pd.DataFrame, class_df: pd.DataFrame, 
                       dec_df: pd.DataFrame) -> None:
    """
    Create a QA workbook for manual verification of extracted data.
    
    This function creates a standardized Excel workbook with multiple sheets
    for quality assurance review of extracted tunnel bunnrensk data.
    
    Args:
        filepath: Path where the Excel workbook will be saved
        summary: Dictionary containing extraction metadata with keys:
            - project: Project name
            - project_code: Project code (e.g., '09_moanetunnelen')
            - extracted_at: ISO timestamp of extraction
            - source_file: Name of source PDF file
            - notat_pages: Page range string for notat (e.g., '2-5')
            - lab_report_pages: Page range string for lab reports
            - lab: Laboratory name
            - lab_reports: Lab report reference numbers
            - sampling_dates: Comma-separated sampling dates
            - samples_count: Number of samples
            - results_count: Number of result rows
            - classifications_count: Number of classifications
            - decisions_count: Number of decisions
        samples_df: DataFrame with sample metadata
        results_df: DataFrame with lab analysis results (long format)
        class_df: DataFrame with tilstandsklasse classifications
        dec_df: DataFrame with handling decisions
    
    Sheets created:
        - Summary: Extraction metadata and file overview
        - Schema Violations: Any values that don't match schema constraints
        - QA Checklist: Standard items to verify with status columns
        - Samples QA: Sample data with QA columns
        - Results QA: Lab results data with QA columns
        - Classifications QA: Classification data with QA columns
        - Decisions QA: Decision data with QA columns
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # Find schema violations
    violations = _find_schema_violations(samples_df, results_df, dec_df)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # =========================================================================
    # SHEET 1: Summary
    # =========================================================================
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Title
    ws_summary['A1'] = f"QA Workbook: {summary.get('project', 'Unknown Project')}"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:D1')
    
    # Extraction info - handle optional keys gracefully
    summary_items = [
        ('Project Code', summary.get('project_code', '')),
        ('Extracted At', summary.get('extracted_at', '')),
        ('Source File', summary.get('source_file', '')),
        ('Notat Pages', summary.get('notat_pages', '')),
        ('Lab Report Pages', summary.get('lab_report_pages', '')),
        ('Lab', summary.get('lab', '')),
        ('Lab Reports', summary.get('lab_reports', '')),
        ('Sampling Dates', summary.get('sampling_dates', '')),
    ]
    
    for i, (label, value) in enumerate(summary_items, start=3):
        ws_summary[f'A{i}'] = label
        ws_summary[f'A{i}'].font = Font(bold=True)
        ws_summary[f'B{i}'] = value
    
    # Data counts
    ws_summary['A13'] = "Data Exports"
    ws_summary['A13'].font = Font(bold=True, size=12)
    
    # Calculate unique parameters for wide format row count
    param_count = len(results_df['parameter'].unique()) if len(results_df) > 0 and 'parameter' in results_df.columns else 0
    
    export_items = [
        ('File', 'Rows', 'Description', 'Remarks', 'Date', 'Name'),
        ('samples.csv', summary.get('samples_count', 0), 'Sample metadata', '', '', ''),
        ('results.csv', summary.get('results_count', 0), 'Lab analysis results (long format)', '', '', ''),
        ('results_wide.csv', param_count, 'Lab results (wide format for review)', '', '', ''),
        ('results_wide.xlsx', param_count, 'Lab results (Excel table)', '', '', ''),
        ('classifications.csv', summary.get('classifications_count', 0), 'Tilstandsklasse per sample', '', '', ''),
        ('decisions.csv', summary.get('decisions_count', 0), 'Handling decisions per sample', '', '', ''),
        ('extraction_summary.csv', 1, 'Extraction metadata', '', '', ''),
    ]
    
    for i, row_data in enumerate(export_items, start=14):
        for j, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=i, column=j, value=value)
            if i == 14:  # Header row
                cell.font = header_font
                cell.fill = header_fill
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 10
    ws_summary.column_dimensions['C'].width = 40
    ws_summary.column_dimensions['D'].width = 30
    ws_summary.column_dimensions['E'].width = 12
    ws_summary.column_dimensions['F'].width = 15
    
    # Add schema violations count to summary
    ws_summary['A23'] = "Schema Validation"
    ws_summary['A23'].font = Font(bold=True, size=12)
    ws_summary['A24'] = "Schema Violations"
    ws_summary['A24'].font = Font(bold=True)
    ws_summary['B24'] = len(violations)
    if len(violations) > 0:
        ws_summary['B24'].fill = error_fill
    else:
        ws_summary['B24'].fill = ok_fill
    ws_summary['C24'] = "See 'Schema Violations' sheet for details" if violations else "All values match schema"
    
    # =========================================================================
    # SHEET 2: Schema Violations
    # =========================================================================
    ws_violations = wb.create_sheet("Schema Violations")
    
    violation_headers = ['Table', 'Sample ID', 'Field', 'Invalid Value', 'Valid Values']
    for j, header in enumerate(violation_headers, start=1):
        cell = ws_violations.cell(row=1, column=j, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    if violations:
        for i, v in enumerate(violations, start=2):
            ws_violations.cell(row=i, column=1, value=v['table']).border = thin_border
            ws_violations.cell(row=i, column=2, value=v['sample_id']).border = thin_border
            ws_violations.cell(row=i, column=3, value=v['field']).border = thin_border
            cell_val = ws_violations.cell(row=i, column=4, value=v['value'])
            cell_val.border = thin_border
            cell_val.fill = error_fill
            ws_violations.cell(row=i, column=5, value=v['valid_values']).border = thin_border
    else:
        ws_violations.cell(row=2, column=1, value="No violations found")
        ws_violations.cell(row=2, column=1).fill = ok_fill
    
    # Set column widths
    ws_violations.column_dimensions['A'].width = 12
    ws_violations.column_dimensions['B'].width = 18
    ws_violations.column_dimensions['C'].width = 15
    ws_violations.column_dimensions['D'].width = 25
    ws_violations.column_dimensions['E'].width = 60
    ws_violations.freeze_panes = 'A2'
    
    # =========================================================================
    # SHEET 3: QA Checklist
    # =========================================================================
    ws_qa = wb.create_sheet("QA Checklist")
    
    checklist_items = [
        ('Category', 'Check Item', 'Status', 'Verified By', 'Date', 'Notes'),
        ('Samples', 'All samples from source document captured', '', '', '', ''),
        ('Samples', 'Sample IDs match lab report numbers', '', '', '', ''),
        ('Samples', 'Sample dates correct', '', '', '', ''),
        ('Samples', 'Profile locations match source', '', '', '', ''),
        ('Results', 'All parameters extracted for each sample', '', '', '', ''),
        ('Results', 'Values match source PDF', '', '', '', ''),
        ('Results', 'Below-limit values (<) correctly flagged', '', '', '', ''),
        ('Results', 'Units correct (mg/kg, %)', '', '', '', ''),
        ('Results', 'Uncertainties captured where available', '', '', '', ''),
        ('Classifications', 'Tilstandsklasse matches source tables', '', '', '', ''),
        ('Classifications', 'Limiting parameters identified correctly', '', '', '', ''),
        ('Classifications', 'Classification basis documented', '', '', '', ''),
        ('Decisions', 'Handling decisions match notat text', '', '', '', ''),
        ('Decisions', 'Destinations correct (deponi, gjenbruk, etc)', '', '', '', ''),
        ('Decisions', 'Responsible parties assigned correctly', '', '', '', ''),
        ('General', 'No duplicate samples', '', '', '', ''),
        ('General', 'No missing critical data', '', '', '', ''),
        ('General', 'Data ready for aggregation', '', '', '', ''),
    ]
    
    for i, row in enumerate(checklist_items, start=1):
        for j, value in enumerate(row, start=1):
            cell = ws_qa.cell(row=i, column=j, value=value)
            cell.border = thin_border
            if i == 1:  # Header
                cell.font = header_font
                cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Set column widths
    ws_qa.column_dimensions['A'].width = 15
    ws_qa.column_dimensions['B'].width = 45
    ws_qa.column_dimensions['C'].width = 12
    ws_qa.column_dimensions['D'].width = 15
    ws_qa.column_dimensions['E'].width = 12
    ws_qa.column_dimensions['F'].width = 40
    
    # Freeze header
    ws_qa.freeze_panes = 'A2'
    
    # =========================================================================
    # SHEET 4: Samples QA
    # =========================================================================
    ws_samples = wb.create_sheet("Samples QA")
    
    # Add QA columns to samples
    samples_qa = samples_df.copy() if len(samples_df) > 0 else pd.DataFrame()
    if len(samples_qa) > 0:
        samples_qa['QA_Status'] = ''
        samples_qa['QA_Notes'] = ''
        
        for i, row in enumerate(dataframe_to_rows(samples_qa, index=False, header=True), start=1):
            for j, value in enumerate(row, start=1):
                cell = ws_samples.cell(row=i, column=j, value=value)
                if i == 1:
                    cell.font = header_font
                    cell.fill = header_fill
        
        ws_samples.freeze_panes = 'A2'
        
        # Auto-width columns
        for col_idx, col in enumerate(samples_qa.columns, start=1):
            ws_samples.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col)) + 2, 12)
    
    # =========================================================================
    # SHEET 5: Results QA
    # =========================================================================
    ws_results = wb.create_sheet("Results QA")
    
    results_qa = results_df.copy() if len(results_df) > 0 else pd.DataFrame()
    if len(results_qa) > 0:
        results_qa['QA_Status'] = ''
        results_qa['QA_Notes'] = ''
        
        for i, row in enumerate(dataframe_to_rows(results_qa, index=False, header=True), start=1):
            for j, value in enumerate(row, start=1):
                cell = ws_results.cell(row=i, column=j, value=value)
                if i == 1:
                    cell.font = header_font
                    cell.fill = header_fill
        
        ws_results.freeze_panes = 'A2'
        
        # Auto-width columns with reasonable max
        for col_idx, col in enumerate(results_qa.columns, start=1):
            ws_results.column_dimensions[get_column_letter(col_idx)].width = min(max(len(str(col)) + 2, 10), 30)
    
    # =========================================================================
    # SHEET 6: Classifications QA
    # =========================================================================
    ws_class = wb.create_sheet("Classifications QA")
    
    class_qa = class_df.copy() if len(class_df) > 0 else pd.DataFrame()
    if len(class_qa) > 0:
        class_qa['QA_Status'] = ''
        class_qa['QA_Notes'] = ''
        
        for i, row in enumerate(dataframe_to_rows(class_qa, index=False, header=True), start=1):
            for j, value in enumerate(row, start=1):
                cell = ws_class.cell(row=i, column=j, value=value)
                if i == 1:
                    cell.font = header_font
                    cell.fill = header_fill
        
        ws_class.freeze_panes = 'A2'
        
        for col_idx, col in enumerate(class_qa.columns, start=1):
            ws_class.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col)) + 2, 15)
    
    # =========================================================================
    # SHEET 7: Decisions QA
    # =========================================================================
    ws_dec = wb.create_sheet("Decisions QA")
    
    dec_qa = dec_df.copy() if len(dec_df) > 0 else pd.DataFrame()
    if len(dec_qa) > 0:
        dec_qa['QA_Status'] = ''
        dec_qa['QA_Notes'] = ''
        
        for i, row in enumerate(dataframe_to_rows(dec_qa, index=False, header=True), start=1):
            for j, value in enumerate(row, start=1):
                cell = ws_dec.cell(row=i, column=j, value=value)
                if i == 1:
                    cell.font = header_font
                    cell.fill = header_fill
        
        ws_dec.freeze_panes = 'A2'
        
        for col_idx, col in enumerate(dec_qa.columns, start=1):
            ws_dec.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col)) + 2, 15)
    
    # Save workbook
    wb.save(filepath)
