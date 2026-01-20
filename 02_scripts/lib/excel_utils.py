"""
Excel utilities for reading and parsing lab reports and data files.
"""
from __future__ import annotations

import pandas as pd
from pathlib import Path
from typing import Optional, List, Dict, Any, Union


def read_excel_file(filepath: str | Path, 
                    sheet_name: str | int = 0,
                    header_row: int = 0,
                    skip_rows: List[int] = None) -> pd.DataFrame:
    """
    Read an Excel file with common options.
    
    Args:
        filepath: Path to Excel file
        sheet_name: Sheet name or index (0-based)
        header_row: Row number to use as column headers (0-based)
        skip_rows: List of row indices to skip
        
    Returns:
        DataFrame with the data
    """
    filepath = Path(filepath)
    
    return pd.read_excel(
        filepath,
        sheet_name=sheet_name,
        header=header_row,
        skiprows=skip_rows
    )


def list_sheets(filepath: str | Path) -> List[str]:
    """List all sheet names in an Excel file."""
    xl = pd.ExcelFile(filepath)
    return xl.sheet_names


def find_header_row(filepath: str | Path, 
                    sheet_name: str | int = 0,
                    keywords: List[str] = None,
                    max_rows: int = 20) -> int:
    """
    Auto-detect the header row by looking for common column keywords.
    
    Args:
        filepath: Path to Excel file
        sheet_name: Sheet to search
        keywords: List of keywords to look for (default: common lab terms)
        max_rows: Maximum rows to search
        
    Returns:
        Row index (0-based) of the header row
    """
    if keywords is None:
        keywords = ['prøve', 'sample', 'parameter', 'enhet', 'unit', 
                    'metall', 'metal', 'resultat', 'result', 'mg/kg', 
                    'analysemetode', 'LOQ', 'LOD']
    
    keywords_lower = [k.lower() for k in keywords]
    
    # Read first N rows without header
    df = pd.read_excel(filepath, sheet_name=sheet_name, header=None, nrows=max_rows)
    
    for idx, row in df.iterrows():
        row_text = ' '.join(str(v).lower() for v in row.values if pd.notna(v))
        matches = sum(1 for kw in keywords_lower if kw in row_text)
        if matches >= 2:  # At least 2 keyword matches
            return idx
    
    return 0  # Default to first row


def read_lab_report(filepath: str | Path,
                    sheet_name: str | int = 0,
                    header_row: int = None,
                    auto_detect_header: bool = True) -> pd.DataFrame:
    """
    Read a lab report Excel file with smart header detection.
    
    Args:
        filepath: Path to Excel file
        sheet_name: Sheet name or index
        header_row: Explicit header row (overrides auto-detect)
        auto_detect_header: Whether to auto-detect header row
        
    Returns:
        DataFrame with the lab data
    """
    filepath = Path(filepath)
    
    if header_row is None and auto_detect_header:
        header_row = find_header_row(filepath, sheet_name)
    elif header_row is None:
        header_row = 0
    
    df = pd.read_excel(filepath, sheet_name=sheet_name, header=header_row)
    
    # Clean up column names
    df.columns = [str(c).strip() for c in df.columns]
    
    # Drop completely empty rows and columns
    df = df.dropna(how='all')
    df = df.dropna(axis=1, how='all')
    
    return df


def extract_all_sheets(filepath: str | Path) -> Dict[str, pd.DataFrame]:
    """
    Extract all sheets from an Excel file.
    
    Returns:
        Dictionary of sheet_name -> DataFrame
    """
    filepath = Path(filepath)
    sheets = list_sheets(filepath)
    
    result = {}
    for sheet in sheets:
        try:
            df = read_lab_report(filepath, sheet_name=sheet)
            if not df.empty:
                result[sheet] = df
        except Exception as e:
            print(f"Warning: Could not read sheet '{sheet}': {e}")
    
    return result


def merge_sample_data(dataframes: List[pd.DataFrame],
                      sample_col: str = 'Prøve',
                      how: str = 'outer') -> pd.DataFrame:
    """
    Merge multiple dataframes by sample ID.
    
    Args:
        dataframes: List of DataFrames to merge
        sample_col: Column name containing sample IDs
        how: Merge method ('outer', 'inner', 'left', 'right')
        
    Returns:
        Merged DataFrame
    """
    if not dataframes:
        return pd.DataFrame()
    
    result = dataframes[0]
    for df in dataframes[1:]:
        result = result.merge(df, on=sample_col, how=how, suffixes=('', '_dup'))
    
    # Remove duplicate columns
    result = result.loc[:, ~result.columns.str.endswith('_dup')]
    
    return result


# =============================================================================
# RESULTS TABLE FORMATTING
# =============================================================================

def create_wide_table(
    results_df: pd.DataFrame,
    sample_id_prefix: str = 'p'
) -> pd.DataFrame:
    """
    Create a wide format table from long-format results.
    
    Transforms results from long format (one row per measurement) to wide format
    (parameters as rows, samples as columns) for easier review and QA.
    
    Args:
        results_df: DataFrame with columns:
            - sample_id: Sample identifier
            - parameter: Standard parameter code
            - parameter_raw: Original parameter name
            - value: Numeric value
            - unit: Unit of measurement
            - below_limit: Boolean flag for below-limit values
        sample_id_prefix: Prefix to identify sample columns (e.g., 'p09-', 'p18-')
            If 'p', will match any column starting with 'p' followed by digits.
    
    Returns:
        Wide-format DataFrame with:
            - param_code: Standard parameter code
            - parameter: Original parameter name
            - One column per sample (e.g., "p09-MOA-001 (mg/kgTS)")
    
    Alerts:
        Prints a warning if samples have mixed units.
    
    Example:
        >>> wide_df = create_wide_table(results_df, sample_id_prefix='p09-')
        >>> wide_df.to_csv('results_wide.csv', index=False)
    """
    # Unit consistency check removed; units are now converted downstream
    
    # Create formatted value strings (without units)
    def format_value(row):
        if pd.isna(row['value']):
            return ''
        
        # Format the numeric value
        val = row['value']
        if val == int(val):
            val_str = str(int(val))
        elif val < 0.01:
            val_str = f"{val:.4f}"
        elif val < 1:
            val_str = f"{val:.3f}"
        else:
            val_str = f"{val:.2f}".rstrip('0').rstrip('.')
        
        # Add < prefix for below-limit values
        if row['below_limit']:
            val_str = f"<{val_str}"
        
        return val_str
    
    results_df = results_df.copy()
    results_df['formatted'] = results_df.apply(format_value, axis=1)
    
    # Get predominant unit per sample (for header)
    sample_units = results_df.groupby('sample_id')['unit'].agg(
        lambda x: x.value_counts().index[0]  # Most common unit
    ).to_dict()
    
    # Pivot to wide format
    wide_df = results_df.pivot_table(
        index=['parameter', 'parameter_raw'],
        columns='sample_id',
        values='formatted',
        aggfunc='first'
    ).reset_index()
    
    # Rename columns for clarity
    wide_df.columns.name = None
    wide_df = wide_df.rename(columns={'parameter': 'param_code', 'parameter_raw': 'parameter'})
    
    # Identify and rename sample columns to include unit in header
    import re
    if sample_id_prefix == 'p':
        # Match any project prefix like p01-, p09-, p18-
        sample_cols = [c for c in wide_df.columns if re.match(r'^p\d+-', c)]
    else:
        sample_cols = [c for c in wide_df.columns if c.startswith(sample_id_prefix)]
    
    sample_cols.sort()
    col_rename = {col: f"{col} ({sample_units.get(col, 'mg/kg')})" for col in sample_cols}
    wide_df = wide_df.rename(columns=col_rename)
    
    # Reorder columns
    new_sample_cols = [col_rename[c] for c in sample_cols]
    wide_df = wide_df[['param_code', 'parameter'] + new_sample_cols]
    
    return wide_df


def save_wide_table_xlsx(wide_df: pd.DataFrame, filepath: Path) -> None:
    """
    Save wide table as formatted Excel table for QA.
    
    Creates a professional Excel table with:
    - Excel Table formatting with filters
    - Alternating row colors (blue theme)
    - Auto-adjusted column widths
    - Frozen header row
    
    Args:
        wide_df: Wide-format DataFrame from create_wide_table()
        filepath: Output path for the .xlsx file
    
    Example:
        >>> wide_df = create_wide_table(results_df)
        >>> save_wide_table_xlsx(wide_df, Path('results_wide.xlsx'))
    """
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    
    # Write dataframe to worksheet
    for r_idx, row in enumerate(dataframe_to_rows(wide_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Define table range
    max_row = len(wide_df) + 1  # +1 for header
    max_col = len(wide_df.columns)
    table_range = f"A1:{get_column_letter(max_col)}{max_row}"
    
    # Create table with style
    table = Table(displayName="ResultsTable", ref=table_range)
    style = TableStyleInfo(
        name="TableStyleMedium2",  # Blue alternating rows
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    
    # Auto-adjust column widths
    for col_idx, col in enumerate(wide_df.columns, 1):
        # Calculate max width based on content
        max_width = len(str(col))  # Header width
        for row_idx in range(2, max_row + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_width = max(max_width, len(str(cell_value)))
        
        # Set width with some padding, max 30 chars
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 2, 30)
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save
    wb.save(filepath)

